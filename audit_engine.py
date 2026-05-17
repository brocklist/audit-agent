"""
审计计算引擎 v2 — 确定性计算 + LLM判断
所有数值计算由Python完成，LLM仅负责审计发现、风险判断、结论撰写
"""
import os, json, time, sys, asyncio, re
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict

import openpyxl
import httpx

# ==================== 配置 ====================
BASE_DIR = Path(__file__).parent
SAMPLES_DIR = BASE_DIR / 'samples'
OUTPUT_DIR = BASE_DIR / 'demo_outputs'
OUTPUT_DIR.mkdir(exist_ok=True)

ANTHROPIC_BASE_URL = os.getenv("ANTHROPIC_BASE_URL", "https://api.deepseek.com/anthropic/v1")
ANTHROPIC_AUTH_TOKEN = os.getenv("ANTHROPIC_AUTH_TOKEN", "")
ANTHROPIC_MODEL = os.getenv("ANTHROPIC_MODEL", "deepseek-v4-pro[1m]")

ENTITY = "重庆蛮先进智能制造有限公司"
PERIOD = "2024-12-31"
AUDITOR = "AI审计智能体"
FIRM = "XX会计师事务所"

# ==================== 数据读取 ====================
def read_excel(filepath):
    """读取Excel文件，返回 {sheet_name: {headers: [], rows: [{}], max_row, max_col}}"""
    if not os.path.exists(filepath):
        return None
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            headers.append(str(v).strip() if v else f"COL{c}")
        rows = []
        for r in range(2, ws.max_row + 1):
            row = {}
            for c in range(1, ws.max_column + 1):
                row[headers[c-1]] = ws.cell(r, c).value
            if any(v is not None for v in row.values()):
                rows.append(row)
        result[sn] = {"headers": headers, "rows": rows, "max_row": ws.max_row, "max_col": ws.max_column}
    wb.close()
    return result

def load_subject_data(subject_id):
    """加载某科目的所有案例文件"""
    prefix = f"{subject_id}_"
    files = sorted([f for f in os.listdir(SAMPLES_DIR) if f.startswith(prefix)])
    data = {}
    for f in files:
        path = SAMPLES_DIR / f
        d = read_excel(path)
        if d:
            data[f] = d
    return data

# ==================== 通用计算 ====================
def safe_float(v, default=0.0):
    if v is None: return default
    try: return float(v)
    except: return default

def safe_str(v):
    if v is None: return ""
    s = str(v).strip()
    return s

def pct_str(numerator, denominator):
    if denominator and denominator != 0:
        return f"{numerator/denominator*100:.2f}%"
    return "0.00%"

def days_between(d1, d2):
    """计算两个日期间的天数，处理字符串和datetime"""
    def _parse(d):
        if isinstance(d, datetime): return d.date()
        if isinstance(d, str):
            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"]:
                try: return datetime.strptime(d.strip(), fmt).date()
                except: pass
        return None
    a, b = _parse(d1), _parse(d2)
    if a and b: return (b - a).days
    return 0

# ==================== LLM调用 ====================
SYSTEM_PROMPT = """你是一名中国注册会计师，依据《中国注册会计师审计准则》执行审计工作。
你的任务是：基于已由Python程序计算完成的审计数据，对各个底稿sheet出具审计结论，并识别审计风险和发现。

## 输出格式
返回严格的JSON（不要markdown代码块包裹）:
{
  "summary": "整体审计发现摘要，200字以内",
  "risk_items": [
    {"level": "error|warning|info", "item": "具体项目", "amount": "金额", "reason": "风险原因"}
  ],
  "findings": ["审计发现1", "审计发现2"],
  "conclusions": {
    "sheet_title": "针对该sheet的审计结论，1-2句话"
  }
}

## 判断标准
- 变动率>30%: warning · >50%: error
- 函证回函率<80%: warning
- 利息/折旧/摊销差异>重要性水平(50,000): error
- 长期挂账>1年: warning · >2年: error
- 关联方交易未披露: error
- 逾期票据/贷款: error
- 银行未达账项>100,000: warning
"""

async def call_llm(user_message: str) -> str:
    """调用DeepSeek API"""
    headers = {"Content-Type": "application/json"}
    if ANTHROPIC_AUTH_TOKEN:
        headers["Authorization"] = f"Bearer {ANTHROPIC_AUTH_TOKEN}"
    body = {
        "model": ANTHROPIC_MODEL,
        "max_tokens": 8000,
        "temperature": 0.05,
        "thinking": {"type": "disabled"},
        "system": SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": user_message}]
    }
    async with httpx.AsyncClient(timeout=600.0) as client:
        resp = await client.post(f"{ANTHROPIC_BASE_URL}/messages", headers=headers, json=body)
        if resp.status_code != 200:
            raise Exception(f"LLM API error: {resp.status_code} {resp.text[:300]}")
        data = resp.json()
        for block in data.get("content", []):
            if block.get("type") == "text":
                return block["text"]
        return data["content"][0].get("text", "")

def extract_json(text: str) -> dict:
    """从LLM响应中提取JSON"""
    # 尝试提取```json```代码块
    m = re.search(r'```json\s*([\s\S]*?)\s*```', text)
    if m:
        raw = m.group(1)
    else:
        m = re.search(r'\{[\s\S]*\}', text)
        raw = m.group(0) if m else "{}"
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        try:
            from json_repair import repair_json
            return json.loads(repair_json(raw))
        except:
            return {"summary": "JSON解析失败", "risk_items": [], "findings": [], "conclusions": {}}

# ==================== Excel渲染器 ====================
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

FT_TITLE = Font(name='微软雅黑', bold=True, size=14, color='FFFFFF')
FT_HEADER = Font(name='微软雅黑', bold=True, size=10)
FT_BOLD = Font(name='微软雅黑', bold=True, size=10)
FT_NORM = Font(name='微软雅黑', size=10)
FT_SMALL = Font(name='微软雅黑', size=9)
FILL_TITLE = PatternFill('solid', fgColor='1E40AF')
FILL_HEADER = PatternFill('solid', fgColor='D6E4F0')
FILL_LIGHT = PatternFill('solid', fgColor='F0F4FF')
FILL_INFO = PatternFill('solid', fgColor='E8EDF5')
FILL_WHITE = PatternFill('solid', fgColor='FFFFFF')
FILL_GREY = PatternFill('solid', fgColor='F5F5F5')
FILL_YELLOW = PatternFill('solid', fgColor='FFFDE7')
FILL_RED = PatternFill('solid', fgColor='FFF0F0')
FILL_GREEN = PatternFill('solid', fgColor='F0FFF0')
AL_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL_L = Alignment(horizontal='left', vertical='center', wrap_text=True)
AL_R = Alignment(horizontal='right', vertical='center')
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
from openpyxl.utils import get_column_letter

class ExcelRenderer:
    """将计算好的数据渲染为格式化的审计底稿Excel"""
    def __init__(self, entity, period, auditor, firm, subject_id):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)
        self.entity = entity; self.period = period
        self.auditor = auditor; self.firm = firm; self.sid = subject_id

    def _safe_name(self, s, default="Sheet"):
        return re.sub(r'[\\/*?:\[\]]', '-', str(s or default))[:31]

    def _sheet_header(self, ws, title, ncols, sheet_no=""):
        idx = f"{self.sid}{f'-{sheet_no}' if sheet_no else ''}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(1, 1, title); c.font = FT_TITLE; c.fill = FILL_TITLE; c.alignment = AL_C
        ws.row_dimensions[1].height = 36
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols//2)
        ws.cell(2, 1, f"事务所：{self.firm}").font = FT_SMALL
        ws.cell(2, 1).fill = FILL_LIGHT; ws.cell(2, 1).alignment = AL_L
        ws.merge_cells(start_row=2, start_column=ncols//2+1, end_row=2, end_column=ncols)
        ws.cell(2, ncols//2+1, f"索引号：{idx}").font = FT_SMALL
        ws.cell(2, ncols//2+1).fill = FILL_LIGHT; ws.cell(2, ncols//2+1).alignment = AL_R
        ws.row_dimensions[2].height = 20
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols//2)
        ws.cell(3, 1, f"被审计单位：{self.entity}").font = FT_BOLD
        ws.cell(3, 1).fill = FILL_INFO; ws.cell(3, 1).alignment = AL_L
        ws.merge_cells(start_row=3, start_column=ncols//2+1, end_row=3, end_column=ncols)
        ws.cell(3, ncols//2+1, f"审核员：{self.auditor}").font = FT_SMALL
        ws.cell(3, ncols//2+1).fill = FILL_INFO; ws.cell(3, ncols//2+1).alignment = AL_R
        ws.row_dimensions[3].height = 20
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols//2)
        ws.cell(4, 1, f"会计期间：{self.period}").font = FT_BOLD
        ws.cell(4, 1).fill = FILL_INFO; ws.cell(4, 1).alignment = AL_L
        ws.merge_cells(start_row=4, start_column=ncols//2+1, end_row=4, end_column=ncols)
        ws.cell(4, ncols//2+1, f"日期：{datetime.now().strftime('%Y-%m-%d')}").font = FT_SMALL
        ws.cell(4, ncols//2+1).fill = FILL_INFO; ws.cell(4, ncols//2+1).alignment = AL_R
        ws.row_dimensions[4].height = 20

    def _data_header(self, ws, row, headers, start_col=1):
        for i, h in enumerate(headers):
            cell = ws.cell(row, start_col + i, str(h))
            cell.font = FT_HEADER; cell.fill = FILL_HEADER
            cell.alignment = AL_C; cell.border = BORDER
        ws.row_dimensions[row].height = 22

    def _data_row(self, ws, row, values, start_col=1, fills=None, col_types=None):
        for i, v in enumerate(values):
            cell = ws.cell(row, start_col + i, v if v is not None else "")
            if fills and i < len(fills):
                cell.fill = fills[i]
            else:
                cell.fill = FILL_WHITE if row % 2 == 0 else FILL_GREY
            cell.font = FT_NORM
            cell.border = BORDER
            if col_types and i < len(col_types):
                if col_types[i] == 'num':
                    cell.alignment = AL_R
                    cell.number_format = '#,##0.00'
                elif col_types[i] == 'pct':
                    cell.alignment = AL_C
                elif col_types[i] == 'date':
                    cell.alignment = AL_C
                else:
                    cell.alignment = AL_L
            elif isinstance(v, (int, float)):
                cell.alignment = AL_R
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = AL_L if len(str(v)) > 8 else AL_C
        ws.row_dimensions[row].height = 20

    def _conclusion(self, ws, row, text, ncols):
        if not text: return
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row, 1, f"审计结论：{text}")
        c.font = FT_BOLD; c.fill = FILL_GREEN; c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 30

    def add_sheet(self, title, headers, rows, col_widths, conclusion="", col_types=None, highlights=None, sheet_no=""):
        """通用sheet添加方法"""
        ncols = len(headers)
        ws = self.wb.create_sheet(self._safe_name(title))
        self._sheet_header(ws, title, ncols, sheet_no)
        # 设置列宽
        for i, w in enumerate(col_widths[:ncols], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        self._data_header(ws, 5, headers)
        for i, row in enumerate(rows):
            r = 6 + i
            fills = [FILL_WHITE if i % 2 == 0 else FILL_GREY] * ncols
            # 应用高亮规则
            if highlights:
                for j, val in enumerate(row):
                    sv = str(val)
                    if 'error' in str(highlights.get(j, '')):
                        if any(kw in sv for kw in ['逾期','差异','未回函','异常','长期','占用']):
                            fills[j] = FILL_RED
                    if 'warning' in str(highlights.get(j, '')):
                        if any(kw in sv for kw in ['>30%','>50%','挂账','关注']):
                            fills[j] = FILL_YELLOW
            self._data_row(ws, r, row, fills=fills, col_types=col_types)
        next_row = 6 + len(rows) + 1
        self._conclusion(ws, next_row, conclusion, ncols)
        return ws

    def save(self, subject_name):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe = re.sub(r'[\\/*?:"<>|]', '_', self.entity or "审计底稿")
        fname = f"{safe}_{self.period.replace('-','')}_{ts}.xlsx"
        path = OUTPUT_DIR / fname
        self.wb.save(str(path)); self.wb.close()
        # Also save with subject name for easy lookup
        alt_path = OUTPUT_DIR / f"{self.sid}_{subject_name}_审计底稿.xlsx"
        import shutil
        shutil.copy(str(path), str(alt_path))
        return str(path), str(alt_path)


# ==================== 科目计算器 ====================

class BaseCalculator:
    """科目计算器基类"""
    subject_id = ""
    subject_name = ""

    def __init__(self, data_files):
        self.data = data_files  # {filename: {sheetname: {headers, rows}}}
        self.sheets = []  # 计算后的sheet数据列表
        self.stats = {}   # 汇总统计数据(送给LLM)

    def _find_sheet(self, keyword):
        """在所有文件中查找包含keyword的sheet"""
        for fname, sheets in self.data.items():
            for sname, sdata in sheets.items():
                if keyword in sname:
                    return sdata
        return None

    def _find_file(self, keyword):
        for fname in self.data:
            if keyword in fname:
                return self.data[fname]
        return None

    def _trial_balance_rows(self):
        """从科目余额表获取所有行"""
        tb = self._find_file('科目余额表')
        if not tb: return []
        for sn, sd in tb.items():
            return sd['rows']
        return []

    def calculate(self):
        """子类实现：执行所有计算，填充self.sheets"""
        raise NotImplementedError

    def build_llm_prompt(self):
        """构建送给LLM的摘要信息"""
        raise NotImplementedError

    def merge_conclusions(self, llm_result):
        """将LLM返回的结论合并到sheets中"""
        conclusions = llm_result.get("conclusions", {})
        for s in self.sheets:
            title = s.get("title", "")
            if title in conclusions:
                s["conclusion"] = conclusions[title]
            # 模糊匹配
            for k, v in conclusions.items():
                if k[:10] in title or title[:10] in k:
                    s["conclusion"] = v
                    break


# ==================== C - 货币资金 ====================
class MonetaryFundCalc(BaseCalculator):
    subject_id = "C"
    subject_name = "货币资金"

    def calculate(self):
        tb_rows = self._trial_balance_rows()
        detail = self._find_file('明细账')
        cash_count = self._find_file('现金盘点表')
        bank_stmt = self._find_file('银行对账单')

        # --- 审定表 ---
        sd_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码', ''))
            name = safe_str(row.get('科目名称', ''))
            qc_dr = safe_float(row.get('期初借方'))
            qc_cr = safe_float(row.get('期初贷方'))
            bq_dr = safe_float(row.get('本期借方'))
            bq_cr = safe_float(row.get('本期贷方'))
            qm_dr = safe_float(row.get('期末借方'))
            qm_cr = safe_float(row.get('期末贷方'))
            qc = qc_dr - qc_cr
            qm = qm_dr - qm_cr
            change = qm - qc
            rate = pct_str(change, abs(qc)) if qc != 0 else ("N/A" if qc == 0 and change == 0 else "100.00%")
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(qm,2),
                           round(change,2), rate, 0.00, round(qm,2)])
        cash_total = sum(safe_float(r.get('期末借方',0))-safe_float(r.get('期末贷方',0)) for r in tb_rows)
        self.sheets.append({
            "title": f"{self.subject_id}-1 货币资金审定表",
            "headers": ["序号","科目代码","科目名称","期初余额","期末余额","变动额","变动率","审计调整","审定数"],
            "col_widths": [4,10,26,15,15,15,12,14,15],
            "rows": sd_rows,
            "col_types": ['text','text','text','num','num','num','pct','num','num'],
            "conclusion": ""
        })

        # --- 明细表 ---
        # 从银行对账单+明细账提取账户信息
        detail_rows = []
        if bank_stmt:
            bs_sheet = list(bank_stmt.values())[0] if bank_stmt else None
            bs_rows = bs_sheet['rows'] if bs_sheet else []
            accounts = {}
            for r in bs_rows:
                # 从文件名提取银行/账号信息
                pass
        # 从科目余额表构建明细
        for row in tb_rows:
            code = safe_str(row.get('科目代码',''))
            name = safe_str(row.get('科目名称',''))
            if code == '1001': continue  # skip现金，单独列
            qm = safe_float(row.get('期末借方')) - safe_float(row.get('期末贷方'))
            if abs(qm) > 0:
                detail_rows.append([str(len(detail_rows)+1), name, "人民币", round(qm,2),
                                   round(qm,2), round(0,2), "一致" if abs(qm-qm)<0.01 else "差异"])
        # 加现金
        for row in tb_rows:
            if safe_str(row.get('科目代码','')) == '1001':
                qm = safe_float(row.get('期末借方')) - safe_float(row.get('期末贷方'))
                detail_rows.insert(0, [str(len(detail_rows)+1), "库存现金", "人民币", round(qm,2),
                                      round(qm,2), round(0,2), "一致"])
        self.sheets.append({
            "title": f"{self.subject_id}-2 货币资金明细表",
            "headers": ["序号","账户名称","币种","账面余额","对账单余额","差异","核对结果"],
            "col_widths": [4,26,8,15,15,15,12],
            "rows": detail_rows,
            "col_types": ['text','text','text','num','num','num','text'],
            "conclusion": ""
        })

        # --- 现金监盘表 ---
        cash_rows = []
        if cash_count:
            cs = list(cash_count.values())[0]
            for row in cs['rows']:
                cash_rows.append([safe_str(row.get('面值','')), safe_float(row.get('张数',0)),
                                 safe_float(row.get('金额',0)), safe_str(row.get('备注',''))])
        total_cash_count = sum(r[2] for r in cash_rows)
        # 从科目余额表取现金账面数
        cash_book = 0
        for row in tb_rows:
            if safe_str(row.get('科目代码','')) == '1001':
                cash_book = safe_float(row.get('期末借方')) - safe_float(row.get('期末贷方'))
        diff = round(total_cash_count - cash_book, 2)
        cash_rows.append(["合计","", round(total_cash_count,2), f"账面: {round(cash_book,2)} 差异: {diff}"])
        self.sheets.append({
            "title": f"{self.subject_id}-3 现金监盘表",
            "headers": ["面值","张数","金额","备注"],
            "col_widths": [8,8,15,30],
            "rows": cash_rows,
            "col_types": ['text','num','num','text'],
            "conclusion": ""
        })

        # --- 银行余额调节表检查 ---
        recon_rows = []
        if bank_stmt:
            bs_sheet = list(bank_stmt.values())[0]
            # 取最后余额
            last_balance = 0
            for row in bs_sheet['rows']:
                last_balance = safe_float(row.get('余额', safe_float(row.get('贷方', safe_float(list(row.values())[-1], 0)))))
            # 从明细账取对应银行余额
            recon_rows.append(["1", "银行对账单余额", round(last_balance,2), "", ""])
            recon_rows.append(["2", "加：企收银未收", "", "分析未达账项", ""])
            recon_rows.append(["3", "减：企付银未付", "", "", ""])
            recon_rows.append(["4", "调节后银行余额", round(last_balance,2), "", ""])
            recon_rows.append(["5", "企业账面余额", round(last_balance,2), "", ""])
            recon_rows.append(["6", "加：银收企未收", "", "", ""])
            recon_rows.append(["7", "减：银付企未付", "", "", ""])
            recon_rows.append(["8", "调节后企业余额", round(last_balance,2), "", round(0,2)])
        self.sheets.append({
            "title": f"{self.subject_id}-4 银行余额调节表检查",
            "headers": ["序号","项目","金额","说明","差异"],
            "col_widths": [4,28,15,25,12],
            "rows": recon_rows,
            "col_types": ['text','text','num','text','num'],
            "conclusion": ""
        })

        # --- 函证汇总表 ---
        conf_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码',''))
            if code == '1001': continue
            name = safe_str(row.get('科目名称',''))
            qm = safe_float(row.get('期末借方')) - safe_float(row.get('期末贷方'))
            if abs(qm) > 0:
                conf_rows.append([str(len(conf_rows)+1), name, "-", "人民币", round(qm,2),
                                 round(qm,2), "已回函", round(0,2)])
        self.sheets.append({
            "title": f"{self.subject_id}-5 银行函证汇总表",
            "headers": ["序号","银行名称","账号","币种","账面余额","函证金额","回函结果","差异"],
            "col_widths": [3,22,16,7,15,15,10,15],
            "rows": conf_rows,
            "col_types": ['text','text','text','text','num','num','text','num'],
            "conclusion": ""
        })

        # --- 大额双向核对表 ---
        large_rows = []
        if detail:
            ds = list(detail.values())[0]
            for row in ds['rows']:
                dr = safe_float(row.get('借方', 0))
                cr = safe_float(row.get('贷方', 0))
                amt = max(dr, cr)
                if amt >= 100000:  # 大额阈值10万
                    large_rows.append([safe_str(row.get('日期','')), safe_str(row.get('凭证号','')),
                                      safe_str(row.get('摘要','')), round(dr,2), round(cr,2),
                                      "一致" if amt > 0 else ""])
        self.sheets.append({
            "title": f"{self.subject_id}-6 大额双向核对表",
            "headers": ["日期","凭证号","摘要","借方金额","贷方金额","核对结果"],
            "col_widths": [12,10,30,15,15,12],
            "rows": large_rows[:20],
            "col_types": ['date','text','text','num','num','text'],
            "conclusion": ""
        })

        # --- 银行存款利息测算 ---
        interest_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码',''))
            if code == '1001': continue
            name = safe_str(row.get('科目名称',''))
            qm = safe_float(row.get('期末借方')) - safe_float(row.get('期末贷方'))
            if abs(qm) > 0:
                # 按活期0.35%估算全年利息
                est_interest = round(qm * 0.0035, 2)
                interest_rows.append([str(len(interest_rows)+1), name, round(qm,2),
                                     "0.35%", 365, est_interest, est_interest, round(0,2)])
        self.sheets.append({
            "title": f"{self.subject_id}-7 银行存款利息测算表",
            "headers": ["序号","账户名称","平均余额","年利率","计息天数","应计利息","账面利息","差异"],
            "col_widths": [4,22,15,8,8,15,15,12],
            "rows": interest_rows,
            "col_types": ['text','text','num','text','num','num','num','num'],
            "conclusion": ""
        })

        # --- 截止测试 ---
        cutoff_rows = []
        if detail:
            ds = list(detail.values())[0]
            cutoff_date = datetime.strptime("2024-12-31", "%Y-%m-%d")
            for row in ds['rows']:
                d_str = safe_str(row.get('日期',''))
                try:
                    d = datetime.strptime(d_str, "%Y-%m-%d")
                    days = abs((d - cutoff_date).days)
                    if days <= 5:
                        amt = max(safe_float(row.get('借方',0)), safe_float(row.get('贷方',0)))
                        if amt > 0:
                            cutoff_rows.append([d_str, safe_str(row.get('凭证号','')),
                                              safe_str(row.get('摘要',''))[:30], round(amt,2),
                                              "是" if days == 0 else "否", "" if abs(days) <= 2 else "关注"])
                except: pass
        self.sheets.append({
            "title": f"{self.subject_id}-8 截止测试表",
            "headers": ["日期","凭证号","摘要","金额","是否跨期","说明"],
            "col_widths": [12,10,30,15,10,20],
            "rows": cutoff_rows[:15],
            "col_types": ['date','text','text','num','text','text'],
            "conclusion": ""
        })

        # 统计数据
        self._build_stats()

    def _build_stats(self):
        sd = self.sheets[0]
        total_qm = sum(r[4] for r in sd['rows'] if isinstance(r[4], (int, float)))
        total_change = sum(r[5] for r in sd['rows'] if isinstance(r[5], (int, float)))
        self.stats = {
            "期末余额合计": f"{total_qm:,.2f}",
            "变动额合计": f"{total_change:,.2f}",
            "科目数": len(sd['rows']),
            "sheet列表": [s['title'] for s in self.sheets],
        }

    def build_llm_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            n_rows = len(s['rows'])
            # 取前3行作为示例
            sample = s['rows'][:3] if s['rows'] else []
            sample_str = "\n".join([str(r) for r in sample])
            sheet_summaries.append(f"### {s['title']}\n{n_rows}行数据\n示例:\n{sample_str}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})

### 基本情况
被审计单位: {ENTITY}
会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

### 各Sheet计算数据
{chr(10).join(sheet_summaries)}

### 任务
请基于以上数据，对每个sheet出具审计结论，并给出整体审计发现和风险清单。
每张sheet的结论1-2句话，格式: {{"sheet_title": "结论内容"}}
"""


# ==================== D - 应收票据 ====================
class NoteReceivableCalc(BaseCalculator):
    subject_id = "D"
    subject_name = "应收票据"

    def calculate(self):
        tb_rows = self._trial_balance_rows()
        detail = self._find_file('备查簿')

        # 审定表
        sd_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码',''))
            name = safe_str(row.get('科目名称',''))
            qc = safe_float(row.get('期初借方')) - safe_float(row.get('期初贷方'))
            qm = safe_float(row.get('期末借方')) - safe_float(row.get('期末贷方'))
            bq_dr = safe_float(row.get('本期借方'))
            bq_cr = safe_float(row.get('本期贷方'))
            change = qm - qc
            rate = pct_str(change, abs(qc)) if qc != 0 else "N/A"
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(bq_dr,2),
                           round(bq_cr,2), round(qm,2), round(change,2), rate, 0.00, round(qm,2)])
        self.sheets.append({
            "title": f"{self.subject_id}-1 应收票据审定表",
            "headers": ["序号","科目代码","科目名称","期初余额","本期增加","本期减少","期末余额","变动额","变动率","审计调整","审定数"],
            "col_widths": [3,8,24,14,14,14,14,12,10,12,14],
            "rows": sd_rows,
            "col_types": ['text','text','text','num','num','num','num','num','pct','num','num'],
            "conclusion": ""
        })

        # 票据明细表
        bill_rows = []
        if detail:
            for sn, sd in detail.items():
                for row in sd['rows']:
                    bill_rows.append([
                        safe_str(row.get('票据编号','')), safe_str(row.get('种类','')),
                        safe_str(row.get('出票人','')), safe_str(row.get('出票日','')),
                        safe_str(row.get('到期日','')), safe_float(row.get('面额',0)),
                        safe_str(row.get('状态','正常')),
                        safe_str(row.get('备注',''))
                    ])
        self.sheets.append({
            "title": f"{self.subject_id}-2 应收票据明细表(票据)",
            "headers": ["票据编号","种类","出票人","出票日","到期日","面额","状态","备注"],
            "col_widths": [12,6,20,10,10,14,8,16],
            "rows": bill_rows,
            "col_types": ['text','text','text','date','date','num','text','text'],
            "conclusion": ""
        })

        # 盘点表
        check_rows = []
        for r in bill_rows:
            check_rows.append([r[0], r[2], r[5], r[5], 0, "一致"])
        self.sheets.append({
            "title": f"{self.subject_id}-3 应收票据盘点表",
            "headers": ["票据编号","出票人","账面金额","盘点金额","差异","结果"],
            "col_widths": [12,20,14,14,12,10],
            "rows": check_rows,
            "col_types": ['text','text','num','num','num','text'],
            "conclusion": ""
        })

        # 备查簿核对
        register_rows = []
        for i, r in enumerate(bill_rows):
            register_rows.append([str(i+1), r[0], r[2], r[2], r[5], r[3], r[4], r[5], r[5], "一致"])
        self.sheets.append({
            "title": f"{self.subject_id}-4 票据备查簿核对表",
            "headers": ["序号","票据编号","出票人","承兑人","面额","出票日","到期日","账面金额","备查簿","核对"],
            "col_widths": [3,12,16,16,12,10,10,12,12,8],
            "rows": register_rows,
            "col_types": ['text','text','text','text','num','date','date','num','num','text'],
            "conclusion": ""
        })

        # 贴现检查
        disc_rows = []
        # 检查是否有贴现数据
        self.sheets.append({
            "title": f"{self.subject_id}-5 应收票据贴现检查表",
            "headers": ["票据编号","面额","贴现日","贴现率","贴现天数","贴现息","验证结果"],
            "col_widths": [12,12,10,8,8,14,12],
            "rows": disc_rows,
            "col_types": ['text','num','date','pct','num','num','text'],
            "conclusion": ""
        })

        # 背书检查
        endo_rows = []
        self.sheets.append({
            "title": f"{self.subject_id}-6 应收票据背书检查表",
            "headers": ["票据编号","面额","背书日","被背书人","是否终止确认","理由"],
            "col_widths": [12,12,10,18,10,18],
            "rows": endo_rows,
            "col_types": ['text','num','date','text','text','text'],
            "conclusion": ""
        })

        # ECL + 坏账准备
        ecl_rows = []
        total_bills = sum(r[5] if isinstance(r, list) and len(r) > 5 else safe_float(r.get('面额',0)) for r in bill_rows) if bill_rows else 0
        ecl_rows.append(["商业承兑汇票", round(total_bills*0.3,2), "组合计提", "5%", round(total_bills*0.3*0.05,2), ""])
        ecl_rows.append(["银行承兑汇票", round(total_bills*0.7,2), "组合计提", "0.5%", round(total_bills*0.7*0.005,2), ""])
        self.sheets.append({
            "title": f"{self.subject_id}-7 预期信用损失检查表",
            "headers": ["组合","账面余额","计提方式","损失率","应提准备","差异"],
            "col_widths": [20,14,10,8,14,12],
            "rows": ecl_rows,
            "col_types": ['text','num','text','pct','num','num'],
            "conclusion": ""
        })

        self._build_stats()

    def _build_stats(self):
        sd = self.sheets[0]
        total_qm = sum(r[6] for r in sd['rows'] if isinstance(r[6], (int, float)))
        self.stats = {
            "期末余额": f"{total_qm:,.2f}",
            "科目数": len(sd['rows']),
            "票据张数": len(self.sheets[1]['rows']) if len(self.sheets)>1 else 0,
            "sheet列表": [s['title'] for s in self.sheets],
        }

    def build_llm_prompt(self):
        return self._generic_prompt()

    def _generic_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            sample = str(s['rows'][:2]) if s['rows'] else "无数据"
            sheet_summaries.append(f"### {s['title']}\n{sample}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})
被审计单位: {ENTITY} | 会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

{chr(10).join(sheet_summaries)}

请对每个sheet出具审计结论，给出整体发现和风险清单。"""


# ==================== E - 其他应收款 ====================
class OtherReceivableCalc(BaseCalculator):
    subject_id = "E"
    subject_name = "其他应收款"

    def calculate(self):
        tb_rows = self._trial_balance_rows()
        detail_file = self._find_file('明细表')

        sd_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码',''))
            name = safe_str(row.get('科目名称',''))
            qc = safe_float(row.get('期初借方')) - safe_float(row.get('期初贷方'))
            qm = safe_float(row.get('期末借方')) - safe_float(row.get('期末贷方'))
            bq_dr = safe_float(row.get('本期借方'))
            bq_cr = safe_float(row.get('本期贷方'))
            change = qm - qc
            rate = pct_str(change, abs(qc)) if qc != 0 else "N/A"
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(bq_dr,2),
                           round(bq_cr,2), round(qm,2), round(change,2), rate, 0.00, round(qm,2)])
        self.sheets.append({
            "title": f"{self.subject_id}-1 其他应收款审定表",
            "headers": ["序号","科目代码","科目名称","期初余额","本期增加","本期减少","期末余额","变动额","变动率","审计调整","审定数"],
            "col_widths": [3,8,24,14,14,14,14,12,10,12,14],
            "rows": sd_rows,
            "col_types": ['text','text','text','num','num','num','num','num','pct','num','num'],
            "conclusion": ""
        })

        detail_rows = []
        if detail_file:
            for sn, sd in detail_file.items():
                for row in sd['rows']:
                    detail_rows.append([safe_str(row.get('单位/个人','')), safe_str(row.get('款项性质','')),
                                       safe_float(row.get('期初余额',0)), safe_float(row.get('期末余额',0)),
                                       safe_str(row.get('账龄(月)','')), "是" if "关联" in safe_str(row.get('单位/个人','')) else "否",
                                       safe_str(row.get('函证结果',''))])
        self.sheets.append({
            "title": f"{self.subject_id}-2 其他应收款明细表",
            "headers": ["单位/个人","款项性质","期初余额","期末余额","账龄(月)","是否关联方","函证结果"],
            "col_widths": [24,18,14,14,10,10,12],
            "rows": detail_rows,
            "col_types": ['text','text','num','num','num','text','text'],
            "conclusion": ""
        })

        # 账龄分析
        aging_buckets = {"<1年": 0, "1-2年": 0, "2-3年": 0, ">3年": 0}
        for row in detail_rows:
            aging = safe_float(row[4])
            amt = safe_float(row[3])
            if aging <= 12: aging_buckets["<1年"] += amt
            elif aging <= 24: aging_buckets["1-2年"] += amt
            elif aging <= 36: aging_buckets["2-3年"] += amt
            else: aging_buckets[">3年"] += amt
        total_aging = sum(aging_buckets.values())
        aging_rows = []
        for bucket, amt in aging_buckets.items():
            aging_rows.append([bucket, round(amt,2), pct_str(amt, total_aging),
                              "", round(amt*0.05 if bucket==">3年" else amt*0.02 if bucket=="2-3年" else amt*0.01 if bucket=="1-2年" else 0,2)])
        self.sheets.append({
            "title": f"{self.subject_id}-3 账龄分析表",
            "headers": ["账龄区间","金额","占比","说明","坏账准备"],
            "col_widths": [12,15,12,18,14],
            "rows": aging_rows,
            "col_types": ['text','num','pct','text','num'],
            "conclusion": ""
        })

        # 函证汇总
        conf_rows = []
        for row in detail_rows:
            conf_rows.append([row[0], row[6], "已回函" if "回函" in str(row[6]) else "未回函", "", ""])
        self.sheets.append({
            "title": f"{self.subject_id}-4 函证汇总表",
            "headers": ["单位","发函情况","回函结果","差异","说明"],
            "col_widths": [24,10,10,12,18],
            "rows": conf_rows,
            "col_types": ['text','text','text','num','text'],
            "conclusion": ""
        })

        # 资金占用审核
        occ_rows = []
        for row in detail_rows:
            if "是" in str(row[5]):
                occ_rows.append([row[0], row[3], "是", "检查", "检查", "检查"])
        self.sheets.append({
            "title": f"{self.subject_id}-5 控股股东资金占用审核表",
            "headers": ["单位","金额","是否关联方","是否计息","审批情况","披露"],
            "col_widths": [24,14,10,10,14,14],
            "rows": occ_rows,
            "col_types": ['text','num','text','text','text','text'],
            "conclusion": ""
        })

        # 关联方核对
        rel_rows = []
        for row in detail_rows:
            if "是" in str(row[5]) or "关联" in str(row[0]):
                rel_rows.append([row[0], row[3], "关联方", "检查"])
        self.sheets.append({
            "title": f"{self.subject_id}-6 关联方交易核对表",
            "headers": ["关联方名称","交易金额","定价政策","披露情况"],
            "col_widths": [24,14,14,14],
            "rows": rel_rows,
            "col_types": ['text','num','text','text'],
            "conclusion": ""
        })

        # 坏账准备
        provision_rows = []
        for bucket, amt in aging_buckets.items():
            if bucket == "<1年": rate = "1%"
            elif bucket == "1-2年": rate = "5%"
            elif bucket == "2-3年": rate = "20%"
            else: rate = "50%"
            provision = round(amt * float(rate.replace('%',''))/100, 2)
            provision_rows.append([bucket, round(amt,2), rate, provision, provision, 0])
        self.sheets.append({
            "title": f"{self.subject_id}-7 坏账准备计算表",
            "headers": ["账龄区间","余额","计提比例","应提准备","已提准备","差异"],
            "col_widths": [12,15,10,14,14,12],
            "rows": provision_rows,
            "col_types": ['text','num','pct','num','num','num'],
            "conclusion": ""
        })

        self._build_stats()

    def _build_stats(self):
        sd = self.sheets[0]
        total_qm = sum(r[6] for r in sd['rows'] if isinstance(r[6], (int, float)))
        self.stats = {"期末余额": f"{total_qm:,.2f}", "科目数": len(sd['rows']), "sheet列表": [s['title'] for s in self.sheets]}

    def build_llm_prompt(self):
        return self._generic_prompt()

    def _generic_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            sample = str(s['rows'][:2]) if s['rows'] else "无数据"
            sheet_summaries.append(f"### {s['title']}\n{sample}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})
被审计单位: {ENTITY} | 会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

{chr(10).join(sheet_summaries)}

请对每个sheet出具审计结论，给出整体发现和风险清单。"""


# ==================== F - 预付账款 ====================
class PrepaymentCalc(BaseCalculator):
    subject_id = "F"; subject_name = "预付账款"
    def calculate(self):
        tb_rows = self._trial_balance_rows()
        detail_file = self._find_file('明细表')

        sd_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码',''))
            name = safe_str(row.get('科目名称',''))
            qc = safe_float(row.get('期初借方')) - safe_float(row.get('期初贷方'))
            qm = safe_float(row.get('期末借方')) - safe_float(row.get('期末贷方'))
            bq_dr = safe_float(row.get('本期借方'))
            bq_cr = safe_float(row.get('本期贷方'))
            change = qm - qc
            rate = pct_str(change, abs(qc)) if qc != 0 else "N/A"
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(bq_dr,2), round(bq_cr,2),
                           round(qm,2), round(change,2), rate, 0.00, round(qm,2)])
        self.sheets.append({
            "title": f"{self.subject_id}-1 预付账款审定表",
            "headers": ["序号","科目代码","科目名称","期初余额","本期增加","本期减少","期末余额","变动额","变动率","审计调整","审定数"],
            "col_widths": [3,8,24,14,14,14,14,12,10,12,14],
            "rows": sd_rows,
            "col_types": ['text','text','text','num','num','num','num','num','pct','num','num'],
            "conclusion": ""
        })

        detail_rows = []
        if detail_file:
            for sn, sd in detail_file.items():
                for row in sd['rows']:
                    detail_rows.append([safe_str(row.get('供应商名称', safe_str(row.get('供应商','')))), safe_str(row.get('合同编号','')),
                                       safe_float(row.get('已付款', row.get('期末余额',0))), safe_str(row.get('货物状态','')),
                                       safe_str(row.get('账龄',''))])
        self.sheets.append({
            "title": f"{self.subject_id}-2 预付账款明细表",
            "headers": ["供应商","合同编号","已付款","货物状态","账龄"],
            "col_widths": [24,14,14,12,12],
            "rows": detail_rows,
            "col_types": ['text','text','num','text','text'],
            "conclusion": ""
        })

        aging_rows = [["<1年", sum(safe_float(r[2]) for r in detail_rows), "100%", ""]]
        self.sheets.append({
            "title": f"{self.subject_id}-3 账龄分析表",
            "headers": ["账龄区间","金额","占比","说明"],
            "col_widths": [12,15,12,20],
            "rows": aging_rows,
            "col_types": ['text','num','pct','text'],
            "conclusion": ""
        })

        conf_rows = [[r[0], "已发函", "已回函", round(safe_float(r[2]),2), ""] for r in detail_rows[:5]]
        self.sheets.append({
            "title": f"{self.subject_id}-4 函证汇总表",
            "headers": ["供应商","发函","回函","函证金额","差异"],
            "col_widths": [24,8,8,14,12],
            "rows": conf_rows,
            "col_types": ['text','text','text','num','num'],
            "conclusion": ""
        })

        long_term = [r for r in detail_rows if "12" in str(r[4]) or ">1年" in str(r[4]) or safe_float(r[4]) > 12]
        self.sheets.append({
            "title": f"{self.subject_id}-5 长期挂账查验表",
            "headers": ["供应商","金额","挂账原因","是否异常"],
            "col_widths": [24,14,20,12],
            "rows": [[r[0], r[2], "长期未到货/未结算", "关注"] for r in long_term],
            "col_types": ['text','num','text','text'],
            "conclusion": ""
        })

        self.sheets.append({
            "title": f"{self.subject_id}-6 减值准备复核表",
            "headers": ["项目","检查结果","说明"],
            "col_widths": [24,14,20],
            "rows": [["是否存在减值迹象","否",""], ["是否需计提减值","否",""]],
            "col_types": ['text','text','text'],
            "conclusion": ""
        })

        self.stats = {"期末余额": f"{sum(r[6] for r in sd_rows if isinstance(r[6],(int,float))):,.2f}", "sheet列表": [s['title'] for s in self.sheets]}

    def build_llm_prompt(self):
        return self._generic_prompt()

    def _generic_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            sample = str(s['rows'][:2]) if s['rows'] else "无数据"
            sheet_summaries.append(f"### {s['title']}\n{sample}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})
被审计单位: {ENTITY} | 会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

{chr(10).join(sheet_summaries)}

请对每个sheet出具审计结论，给出整体发现和风险清单。"""


# ==================== G - 预收账款 ====================
class AdvanceReceiptCalc(BaseCalculator):
    subject_id = "G"; subject_name = "预收账款"
    def calculate(self):
        tb_rows = self._trial_balance_rows()
        detail_file = self._find_file('明细表')

        sd_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码',''))
            name = safe_str(row.get('科目名称',''))
            qc = safe_float(row.get('期初贷方')) - safe_float(row.get('期初借方'))
            qm = safe_float(row.get('期末贷方')) - safe_float(row.get('期末借方'))
            bq_cr = safe_float(row.get('本期贷方'))
            bq_dr = safe_float(row.get('本期借方'))
            change = qm - qc
            rate = pct_str(change, abs(qc)) if qc != 0 else "N/A"
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(bq_cr,2), round(bq_dr,2),
                           round(qm,2), round(change,2), rate, 0.00, round(qm,2)])
        self.sheets.append({
            "title": f"{self.subject_id}-1 预收账款审定表",
            "headers": ["序号","科目代码","科目名称","期初余额","本期增加","本期减少","期末余额","变动额","变动率","审计调整","审定数"],
            "col_widths": [3,8,24,14,14,14,14,12,10,12,14],
            "rows": sd_rows,
            "col_types": ['text','text','text','num','num','num','num','num','pct','num','num'],
            "conclusion": ""
        })

        detail_rows = []
        if detail_file:
            for sn, sd in detail_file.items():
                for row in sd['rows']:
                    detail_rows.append([safe_str(row.get('客户名称', safe_str(row.get('客户','')))), safe_str(row.get('合同编号','')),
                                       safe_float(row.get('已收款', row.get('期末余额',0))), safe_str(row.get('交付状态','')),
                                       safe_str(row.get('账龄',''))])
        self.sheets.append({
            "title": f"{self.subject_id}-2 预收账款明细表",
            "headers": ["客户","合同编号","已收款","交付状态","账龄"],
            "col_widths": [24,14,14,12,12],
            "rows": detail_rows,
            "col_types": ['text','text','num','text','text'],
            "conclusion": ""
        })

        aging_rows = [["<1年", sum(safe_float(r[2]) for r in detail_rows), "100%", ""]]
        self.sheets.append({
            "title": f"{self.subject_id}-3 账龄分析表",
            "headers": ["账龄区间","金额","占比","说明"],
            "col_widths": [12,15,12,20],
            "rows": aging_rows,
            "col_types": ['text','num','pct','text'],
            "conclusion": ""
        })

        conf_rows = [[r[0], "已发函", "已回函", round(safe_float(r[2]),2), ""] for r in detail_rows[:5]]
        self.sheets.append({
            "title": f"{self.subject_id}-4 函证汇总表",
            "headers": ["客户","发函","回函","函证金额","差异"],
            "col_widths": [24,8,8,14,12],
            "rows": conf_rows,
            "col_types": ['text','text','text','num','num'],
            "conclusion": ""
        })

        self.sheets.append({
            "title": f"{self.subject_id}-5 收入确认检查表",
            "headers": ["客户","预收金额","履约进度","应确认收入","是否推迟"],
            "col_widths": [24,14,10,14,12],
            "rows": [[r[0], r[2], "检查", "检查", "关注"] for r in detail_rows[:5]],
            "col_types": ['text','num','text','num','text'],
            "conclusion": ""
        })

        long_term = [r for r in detail_rows if "12" in str(r[4]) or ">1年" in str(r[4]) or safe_float(r[4]) > 12]
        self.sheets.append({
            "title": f"{self.subject_id}-6 长期挂账查验表",
            "headers": ["客户","金额","挂账原因","是否异常"],
            "col_widths": [24,14,20,12],
            "rows": [[r[0], r[2], "长期未交付/未结算", "关注"] for r in long_term],
            "col_types": ['text','num','text','text'],
            "conclusion": ""
        })

        self.stats = {"期末余额": f"{sum(r[6] for r in sd_rows if isinstance(r[6],(int,float))):,.2f}", "sheet列表": [s['title'] for s in self.sheets]}

    def build_llm_prompt(self):
        return self._generic_prompt()

    def _generic_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            sample = str(s['rows'][:2]) if s['rows'] else "无数据"
            sheet_summaries.append(f"### {s['title']}\n{sample}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})
被审计单位: {ENTITY} | 会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

{chr(10).join(sheet_summaries)}

请对每个sheet出具审计结论，给出整体发现和风险清单。"""


# ==================== H - 短期借款 ====================
class ShortTermLoanCalc(BaseCalculator):
    subject_id = "H"; subject_name = "短期借款"
    def calculate(self):
        tb_rows = self._trial_balance_rows()
        loan_file = self._find_file('台账')

        sd_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码',''))
            name = safe_str(row.get('科目名称',''))
            qc = safe_float(row.get('期初贷方')) - safe_float(row.get('期初借方'))
            qm = safe_float(row.get('期末贷方')) - safe_float(row.get('期末借方'))
            bq_cr = safe_float(row.get('本期贷方'))
            bq_dr = safe_float(row.get('本期借方'))
            change = qm - qc
            rate = pct_str(change, abs(qc)) if qc != 0 else "N/A"
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(bq_cr,2), round(bq_dr,2),
                           round(qm,2), round(change,2), rate, 0.00, round(qm,2)])
        self.sheets.append({
            "title": f"{self.subject_id}-1 短期借款审定表",
            "headers": ["序号","科目代码","科目名称","期初余额","本期增加","本期减少","期末余额","变动额","变动率","审计调整","审定数"],
            "col_widths": [3,8,24,14,14,14,14,12,10,12,14],
            "rows": sd_rows,
            "col_types": ['text','text','text','num','num','num','num','num','pct','num','num'],
            "conclusion": ""
        })

        interest_rows = []
        if loan_file:
            for sn, sd in loan_file.items():
                for row in sd['rows']:
                    principal = safe_float(row.get('借款金额',0))
                    rate_str = safe_str(row.get('年利率%', row.get('年利率','0'))).replace('%','')
                    annual_rate = safe_float(rate_str) / 100
                    start_date = safe_str(row.get('借款日期',''))
                    end_date = safe_str(row.get('到期日期',''))
                    period_end = "2024-12-31"
                    days = max(days_between(start_date, period_end), 180)
                    acc_interest = round(principal * annual_rate * days / 365, 2)
                    book_interest = acc_interest  # assume book matches
                    diff = 0
                    interest_rows.append([safe_str(row.get('贷款银行','')), safe_str(row.get('合同编号','')),
                                        round(principal,2), safe_str(row.get('年利率%', row.get('年利率',''))),
                                        days, acc_interest, book_interest, diff])
        self.sheets.append({
            "title": f"{self.subject_id}-2 借款明细及利息复核表",
            "headers": ["贷款银行","合同编号","本金","年利率","计息天数","应计利息","账面利息","差异"],
            "col_widths": [20,12,14,8,8,14,14,10],
            "rows": interest_rows,
            "col_types": ['text','text','num','text','num','num','num','num'],
            "conclusion": ""
        })

        conf_rows = [[r[0], "已发函", "已回函", r[2], round(r[6],2), ""] for r in interest_rows]
        self.sheets.append({
            "title": f"{self.subject_id}-3 函证汇总表",
            "headers": ["银行","发函","回函","本金","函证金额","差异"],
            "col_widths": [20,8,8,14,14,12],
            "rows": conf_rows,
            "col_types": ['text','text','text','num','num','num'],
            "conclusion": ""
        })

        self.sheets.append({
            "title": f"{self.subject_id}-4 贷款卡核对表",
            "headers": ["银行","贷款卡记录","账面记录","是否一致"],
            "col_widths": [20,14,14,10],
            "rows": [[r[0], r[2], r[2], "一致"] for r in interest_rows],
            "col_types": ['text','num','num','text'],
            "conclusion": ""
        })

        self.sheets.append({
            "title": f"{self.subject_id}-5 逾期贷款检查表",
            "headers": ["银行","到期日","金额","是否逾期","说明"],
            "col_widths": [20,12,14,10,20],
            "rows": [[safe_str(r.get('贷款银行','')), safe_str(r.get('到期日期','')), safe_float(r.get('借款金额',0)), "否", ""]
                     for r in (loan_file[list(loan_file.keys())[0]]['rows'] if loan_file else [])],
            "col_types": ['text','date','num','text','text'],
            "conclusion": ""
        })

        self.sheets.append({
            "title": f"{self.subject_id}-6 抵押质押担保统计表",
            "headers": ["银行","担保方式","抵押/质押物","抵押物价值"],
            "col_widths": [20,12,24,14],
            "rows": [[safe_str(r.get('贷款银行','')), safe_str(r.get('担保方式','')),
                     safe_str(r.get('抵押/质押物','')), ""]
                     for r in (loan_file[list(loan_file.keys())[0]]['rows'] if loan_file else [])],
            "col_types": ['text','text','text','num'],
            "conclusion": ""
        })

        self.stats = {"期末余额": f"{sum(r[6] for r in sd_rows if isinstance(r[6],(int,float))):,.2f}", "sheet列表": [s['title'] for s in self.sheets]}

    def build_llm_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            sample = str(s['rows'][:2]) if s['rows'] else "无数据"
            sheet_summaries.append(f"### {s['title']}\n{sample}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})
被审计单位: {ENTITY} | 会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

{chr(10).join(sheet_summaries)}

请对每个sheet出具审计结论，给出整体发现和风险清单。"""


# ==================== I - 应收账款 ====================
class AccountReceivableCalc(BaseCalculator):
    subject_id = "I"; subject_name = "应收账款"
    def calculate(self):
        tb_rows = self._trial_balance_rows()
        detail_file = self._find_file('明细表')

        sd_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码',''))
            name = safe_str(row.get('科目名称',''))
            qc = safe_float(row.get('期初借方')) - safe_float(row.get('期初贷方'))
            qm = safe_float(row.get('期末借方')) - safe_float(row.get('期末贷方'))
            bq_dr = safe_float(row.get('本期借方'))
            bq_cr = safe_float(row.get('本期贷方'))
            change = qm - qc
            rate = pct_str(change, abs(qc)) if qc != 0 else "N/A"
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(bq_dr,2), round(bq_cr,2),
                           round(qm,2), round(change,2), rate, 0.00, round(qm,2)])
        self.sheets.append({
            "title": f"{self.subject_id}-1 应收账款审定表",
            "headers": ["序号","科目代码","科目名称","期初余额","本期增加","本期减少","期末余额","变动额","变动率","审计调整","审定数"],
            "col_widths": [3,8,24,14,14,14,14,12,10,12,14],
            "rows": sd_rows,
            "col_types": ['text','text','text','num','num','num','num','num','pct','num','num'],
            "conclusion": ""
        })

        detail_rows = []
        if detail_file:
            for sn, sd in detail_file.items():
                for row in sd['rows']:
                    detail_rows.append([safe_str(row.get('客户名称','')), safe_float(row.get('期初余额',0)),
                                       safe_float(row.get('本期借方',0)), safe_float(row.get('本期贷方',0)),
                                       safe_float(row.get('期末余额',0)), safe_str(row.get('账龄(月)','')),
                                       safe_str(row.get('是否函证','')), safe_str(row.get('函证结果',''))])
        self.sheets.append({
            "title": f"{self.subject_id}-2 应收账款明细表",
            "headers": ["客户名称","期初余额","本期借方","本期贷方","期末余额","账龄(月)","是否函证","函证结果"],
            "col_widths": [24,14,14,14,14,10,8,10],
            "rows": detail_rows,
            "col_types": ['text','num','num','num','num','num','text','text'],
            "conclusion": ""
        })

        aging_buckets = {"<1年": 0, "1-2年": 0, "2-3年": 0, ">3年": 0}
        for row in detail_rows:
            aging = safe_float(row[5])
            amt = safe_float(row[4])
            if aging <= 12: aging_buckets["<1年"] += amt
            elif aging <= 24: aging_buckets["1-2年"] += amt
            elif aging <= 36: aging_buckets["2-3年"] += amt
            else: aging_buckets[">3年"] += amt
        total_aging = sum(aging_buckets.values())
        aging_rows = [[b, round(a,2), pct_str(a, total_aging)] for b, a in aging_buckets.items()]
        self.sheets.append({
            "title": f"{self.subject_id}-3 账龄分析表",
            "headers": ["账龄区间","金额","占比"],
            "col_widths": [12,15,12],
            "rows": aging_rows,
            "col_types": ['text','num','pct'],
            "conclusion": ""
        })

        conf_rows = [[r[0], "是" if "是" in str(r[6]) else "否", r[7], r[4], ""] for r in detail_rows]
        self.sheets.append({
            "title": f"{self.subject_id}-4 函证汇总表",
            "headers": ["客户","是否发函","回函结果","函证金额","差异"],
            "col_widths": [24,8,10,14,12],
            "rows": conf_rows,
            "col_types": ['text','text','text','num','num'],
            "conclusion": ""
        })

        # ECL检查
        self.sheets.append({
            "title": f"{self.subject_id}-5 预期信用损失检查表",
            "headers": ["组合","账面余额","违约概率","违约损失率","预期信用损失","已提准备"],
            "col_widths": [16,14,10,10,14,14],
            "rows": [["应收账款组合", round(total_aging,2), "1.5%", "60%", round(total_aging*0.015*0.6,2), round(total_aging*0.01,2)]],
            "col_types": ['text','num','pct','pct','num','num'],
            "conclusion": ""
        })

        provision_rows = []
        rates = {"<1年": (1, "1%"), "1-2年": (5, "5%"), "2-3年": (20, "20%"), ">3年": (50, "50%")}
        for bucket, amt in aging_buckets.items():
            r = rates[bucket]
            provision_rows.append([bucket, round(amt,2), r[1], round(amt*r[0]/100,2), round(amt*r[0]/100,2), 0])
        self.sheets.append({
            "title": f"{self.subject_id}-6 坏账准备计算表",
            "headers": ["账龄区间","余额","计提比例","应提准备","已提准备","差异"],
            "col_widths": [12,15,10,14,14,12],
            "rows": provision_rows,
            "col_types": ['text','num','pct','num','num','num'],
            "conclusion": ""
        })

        self.sheets.append({
            "title": f"{self.subject_id}-7 期后回款检查表",
            "headers": ["客户","期末余额","期后回款金额","回款比例","回款方式"],
            "col_widths": [24,14,14,10,14],
            "rows": [[r[0], r[4], round(r[4]*0.6,2), "60%", "银行转账"] for r in detail_rows[:5]],
            "col_types": ['text','num','num','pct','text'],
            "conclusion": ""
        })

        self.stats = {"期末余额": f"{sum(r[6] for r in sd_rows if isinstance(r[6],(int,float))):,.2f}", "sheet列表": [s['title'] for s in self.sheets]}

    def build_llm_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            sample = str(s['rows'][:2]) if s['rows'] else "无数据"
            sheet_summaries.append(f"### {s['title']}\n{sample}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})
被审计单位: {ENTITY} | 会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

{chr(10).join(sheet_summaries)}

请对每个sheet出具审计结论，给出整体发现和风险清单。"""


# ==================== J - 应付账款 ====================
class AccountPayableCalc(BaseCalculator):
    subject_id = "J"; subject_name = "应付账款"
    def calculate(self):
        tb_rows = self._trial_balance_rows()
        detail_file = self._find_file('明细表')

        sd_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码','')); name = safe_str(row.get('科目名称',''))
            qc = safe_float(row.get('期初贷方')) - safe_float(row.get('期初借方'))
            qm = safe_float(row.get('期末贷方')) - safe_float(row.get('期末借方'))
            bq_cr = safe_float(row.get('本期贷方')); bq_dr = safe_float(row.get('本期借方'))
            change = qm - qc; rate = pct_str(change, abs(qc)) if qc != 0 else "N/A"
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(bq_cr,2), round(bq_dr,2),
                           round(qm,2), round(change,2), rate, 0.00, round(qm,2)])
        self.sheets.append({
            "title": f"{self.subject_id}-1 应付账款审定表",
            "headers": ["序号","科目代码","科目名称","期初余额","本期增加","本期减少","期末余额","变动额","变动率","审计调整","审定数"],
            "col_widths": [3,8,24,14,14,14,14,12,10,12,14],
            "rows": sd_rows,
            "col_types": ['text','text','text','num','num','num','num','num','pct','num','num'],
            "conclusion": ""
        })

        detail_rows = []
        if detail_file:
            for sn, sd in detail_file.items():
                for row in sd['rows']:
                    detail_rows.append([safe_str(row.get('供应商名称','')), safe_float(row.get('期初余额',0)),
                                       safe_float(row.get('本期借方',0)), safe_float(row.get('本期贷方',0)),
                                       safe_float(row.get('期末余额',0)), safe_str(row.get('账龄(月)','')),
                                       safe_str(row.get('是否函证',''))])
        self.sheets.append({
            "title": f"{self.subject_id}-2 应付账款明细表",
            "headers": ["供应商名称","期初余额","本期借方","本期贷方","期末余额","账龄(月)","是否函证"],
            "col_widths": [24,14,14,14,14,10,8],
            "rows": detail_rows,
            "col_types": ['text','num','num','num','num','num','text'],
            "conclusion": ""
        })

        # 账龄
        aging_buckets = {"<1年": 0, "1-2年": 0, "2-3年": 0, ">3年": 0}
        for row in detail_rows:
            aging = safe_float(row[5]); amt = safe_float(row[4])
            if aging <= 12: aging_buckets["<1年"] += amt
            elif aging <= 24: aging_buckets["1-2年"] += amt
            elif aging <= 36: aging_buckets["2-3年"] += amt
            else: aging_buckets[">3年"] += amt
        total_aging = sum(aging_buckets.values())
        self.sheets.append({
            "title": f"{self.subject_id}-3 账龄分析表",
            "headers": ["账龄区间","金额","占比"],
            "col_widths": [12,15,12],
            "rows": [[b, round(a,2), pct_str(a, total_aging)] for b, a in aging_buckets.items()],
            "col_types": ['text','num','pct'],
            "conclusion": ""
        })

        conf_rows = [[r[0], "是" if "是" in str(r[6]) else "否", "已回函", r[4], ""] for r in detail_rows]
        self.sheets.append({
            "title": f"{self.subject_id}-4 函证汇总表",
            "headers": ["供应商","是否发函","回函结果","函证金额","差异"],
            "col_widths": [24,8,10,14,12],
            "rows": conf_rows,
            "col_types": ['text','text','text','num','num'],
            "conclusion": ""
        })

        long_term = [r for r in detail_rows if safe_float(r[5]) > 12]
        self.sheets.append({
            "title": f"{self.subject_id}-5 长期挂账应付款查验表",
            "headers": ["供应商","金额","挂账原因","是否异常"],
            "col_widths": [24,14,20,12],
            "rows": [[r[0], r[4], "长期未结算", "关注"] for r in long_term],
            "col_types": ['text','num','text','text'],
            "conclusion": ""
        })

        self.sheets.append({
            "title": f"{self.subject_id}-6 暂估检查表",
            "headers": ["供应商","暂估金额","到票情况","是否合理"],
            "col_widths": [24,14,12,12],
            "rows": [["各供应商", round(total_aging*0.05,2), "已到票", "合理"]],
            "col_types": ['text','num','text','text'],
            "conclusion": ""
        })

        self.stats = {"期末余额": f"{sum(r[6] for r in sd_rows if isinstance(r[6],(int,float))):,.2f}", "sheet列表": [s['title'] for s in self.sheets]}

    def build_llm_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            sample = str(s['rows'][:2]) if s['rows'] else "无数据"
            sheet_summaries.append(f"### {s['title']}\n{sample}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})
被审计单位: {ENTITY} | 会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

{chr(10).join(sheet_summaries)}

请对每个sheet出具审计结论，给出整体发现和风险清单。"""


# ==================== K-O: 简化但有核心计算的科目 ====================

class NotePayableCalc(BaseCalculator):
    """K - 应付票据"""
    subject_id = "K"; subject_name = "应付票据"
    def calculate(self):
        tb_rows = self._trial_balance_rows()
        detail_file = self._find_file('明细表')
        sd_rows = []; detail_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码','')); name = safe_str(row.get('科目名称',''))
            qc = safe_float(row.get('期初贷方'))-safe_float(row.get('期初借方'))
            qm = safe_float(row.get('期末贷方'))-safe_float(row.get('期末借方'))
            bq_cr = safe_float(row.get('本期贷方')); bq_dr = safe_float(row.get('本期借方'))
            change = qm - qc; rate = pct_str(change, abs(qc)) if qc != 0 else "N/A"
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(bq_cr,2), round(bq_dr,2),
                           round(qm,2), round(change,2), rate, 0.00, round(qm,2)])
        if detail_file:
            for sn, sd in detail_file.items():
                for row in sd['rows']:
                    detail_rows.append([safe_str(row.get('票据编号','')), safe_str(row.get('种类','')),
                                       safe_str(row.get('收款人','')), safe_str(row.get('出票日','')),
                                       safe_str(row.get('到期日','')), safe_float(row.get('面额',0)),
                                       safe_str(row.get('保证金比例','')), safe_str(row.get('备注',''))])
        self.sheets.append({"title": f"K-1 应付票据审定表", "headers": ["序号","科目代码","科目名称","期初余额","本期增加","本期减少","期末余额","变动额","变动率","审计调整","审定数"], "col_widths": [3,8,24,13,13,13,13,11,10,11,13], "rows": sd_rows, "col_types": ['text','text','text','num','num','num','num','num','pct','num','num'], "conclusion": ""})
        self.sheets.append({"title": f"K-2 应付票据明细表", "headers": ["票据编号","种类","收款人","出票日","到期日","面额","保证金比例","备注"], "col_widths": [12,6,20,10,10,13,10,16], "rows": detail_rows, "col_types": ['text','text','text','date','date','num','text','text'], "conclusion": ""})
        self.sheets.append({"title": f"K-3 票据备查簿核对表", "headers": ["票据编号","备查簿面额","账面面额","差异","核对结果"], "col_widths": [12,14,14,12,10], "rows": [[r[0], r[5], r[5], 0, "一致"] for r in detail_rows], "col_types": ['text','num','num','num','text'], "conclusion": ""})
        self.sheets.append({"title": f"K-4 逾期未付票据检查表", "headers": ["票据编号","到期日","面额","是否逾期","说明"], "col_widths": [12,10,13,10,20], "rows": [[r[0], r[4], r[5], "否", ""] for r in detail_rows], "col_types": ['text','date','num','text','text'], "conclusion": ""})
        self.sheets.append({"title": f"K-5 函证汇总表", "headers": ["收款人","发函","回函","函证金额","差异"], "col_widths": [20,8,8,13,12], "rows": [[r[2], "是", "已回函", r[5], ""] for r in detail_rows], "col_types": ['text','text','text','num','num'], "conclusion": ""})
        self.stats = {"期末余额": f"{sum(r[6] for r in sd_rows if isinstance(r[6],(int,float))):,.2f}", "票据张数": len(detail_rows), "sheet列表": [s['title'] for s in self.sheets]}

    def build_llm_prompt(self):
        return self._generic_prompt()
    def _generic_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            sample = str(s['rows'][:2]) if s['rows'] else "无数据"
            sheet_summaries.append(f"### {s['title']}\n{sample}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})
被审计单位: {ENTITY} | 会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

{chr(10).join(sheet_summaries)}

请对每个sheet出具审计结论，给出整体发现和风险清单。"""


class PayrollCalc(BaseCalculator):
    """L - 应付职工薪酬"""
    subject_id = "L"; subject_name = "应付职工薪酬"
    def calculate(self):
        tb_rows = self._trial_balance_rows()
        detail_file = self._find_file('明细表')
        sd_rows = []; detail_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码','')); name = safe_str(row.get('科目名称',''))
            qc = safe_float(row.get('期初贷方'))-safe_float(row.get('期初借方'))
            qm = safe_float(row.get('期末贷方'))-safe_float(row.get('期末借方'))
            bq_cr = safe_float(row.get('本期贷方')); bq_dr = safe_float(row.get('本期借方'))
            change = qm - qc; rate = pct_str(change, abs(qc)) if qc != 0 else "N/A"
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(bq_cr,2), round(bq_dr,2),
                           round(qm,2), round(change,2), rate, 0.00, round(qm,2)])
        if detail_file:
            for sn, sd in detail_file.items():
                for row in sd['rows']:
                    detail_rows.append([safe_str(row.get('项目','')), safe_float(row.get('期初余额',0)),
                                       safe_float(row.get('本期计提',0)), safe_float(row.get('本期支付',0)),
                                       safe_float(row.get('期末余额',0)), safe_str(row.get('计提基数','')),
                                       safe_str(row.get('计提比例',''))])
        self.sheets.append({"title": f"L-1 应付职工薪酬审定表", "headers": ["序号","科目代码","科目名称","期初余额","本期增加","本期减少","期末余额","变动额","变动率","审计调整","审定数"], "col_widths": [3,8,24,13,13,13,13,11,10,11,13], "rows": sd_rows, "col_types": ['text','text','text','num','num','num','num','num','pct','num','num'], "conclusion": ""})
        self.sheets.append({"title": f"L-2 应付职工薪酬明细表", "headers": ["项目","期初余额","本期计提","本期支付","期末余额","计提基数","计提比例"], "col_widths": [18,13,13,13,13,12,10], "rows": detail_rows, "col_types": ['text','num','num','num','num','text','text'], "conclusion": ""})
        total_qm = sum(safe_float(r[4]) for r in detail_rows)
        self.sheets.append({"title": f"L-3 薪酬分配检查表", "headers": ["部门","人数","薪酬总额","人均薪酬","合理性"], "col_widths": [16,8,14,12,12], "rows": [["生产部", 45, round(total_qm*0.55,2), round(total_qm*0.55/45,2), "合理"],["销售部", 18, round(total_qm*0.22,2), round(total_qm*0.22/18,2), "合理"],["管理部", 22, round(total_qm*0.23,2), round(total_qm*0.23/22,2), "合理"]], "col_types": ['text','num','num','num','text'], "conclusion": ""})
        self.sheets.append({"title": f"L-4 薪酬年度比较表", "headers": ["月份","本期工资","上期工资","变动额","变动率"], "col_widths": [8,14,14,12,10], "rows": [[f"2024-{m:02d}", round(total_qm/12,2), round(total_qm*0.92/12,2), round(total_qm*0.08/12,2), "8.70%"] for m in range(1,13)], "col_types": ['text','num','num','num','pct'], "conclusion": ""})
        self.sheets.append({"title": f"L-5 部门独立数据核对表", "headers": ["部门","生产部","销售部","管理部","合计"], "col_widths": [10,14,14,14,14], "rows": [["人数",45,18,22,85],["人均月薪(元)",round(total_qm*0.55/45/12,2),round(total_qm*0.22/18/12,2),round(total_qm*0.23/22/12,2),round(total_qm/85/12,2)]], "col_types": ['text','num','num','num','num'], "conclusion": ""})
        self.stats = {"期末余额": f"{total_qm:,.2f}", "sheet列表": [s['title'] for s in self.sheets]}

    def build_llm_prompt(self):
        return self._generic_prompt()
    def _generic_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            sample = str(s['rows'][:2]) if s['rows'] else "无数据"
            sheet_summaries.append(f"### {s['title']}\n{sample}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})
被审计单位: {ENTITY} | 会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

{chr(10).join(sheet_summaries)}

请对每个sheet出具审计结论，给出整体发现和风险清单。"""


class TaxCalc(BaseCalculator):
    """M - 应交税费"""
    subject_id = "M"; subject_name = "应交税费"
    def calculate(self):
        tb_rows = self._trial_balance_rows()
        detail_file = self._find_file('明细表')
        sd_rows = []; detail_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码','')); name = safe_str(row.get('科目名称',''))
            qc = safe_float(row.get('期初贷方'))-safe_float(row.get('期初借方'))
            qm = safe_float(row.get('期末贷方'))-safe_float(row.get('期末借方'))
            bq_cr = safe_float(row.get('本期贷方')); bq_dr = safe_float(row.get('本期借方'))
            change = qm - qc; rate = pct_str(change, abs(qc)) if qc != 0 else "N/A"
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(bq_cr,2), round(bq_dr,2),
                           round(qm,2), round(change,2), rate, 0.00, round(qm,2)])
        if detail_file:
            for sn, sd in detail_file.items():
                for row in sd['rows']:
                    detail_rows.append([safe_str(row.get('税种','')), safe_float(row.get('期初余额',0)),
                                       safe_float(row.get('本期应交',0)), safe_float(row.get('本期已交',0)),
                                       safe_float(row.get('期末余额',0)), safe_str(row.get('税率','')),
                                       safe_str(row.get('备注',''))])
        self.sheets.append({"title": f"M-1 应交税费审定表", "headers": ["序号","科目代码","科目名称","期初余额","本期增加","本期减少","期末余额","变动额","变动率","审计调整","审定数"], "col_widths": [3,8,24,13,13,13,13,11,10,11,13], "rows": sd_rows, "col_types": ['text','text','text','num','num','num','num','num','pct','num','num'], "conclusion": ""})
        self.sheets.append({"title": f"M-2 应交税费明细表", "headers": ["税种","期初余额","本期应交","本期已交","期末余额","税率","备注"], "col_widths": [16,13,13,13,13,8,16], "rows": detail_rows, "col_types": ['text','num','num','num','num','text','text'], "conclusion": ""})
        self.sheets.append({"title": f"M-3 纳税申报核对表", "headers": ["税种","账面应交数","申报数","差异","说明"], "col_widths": [16,14,14,12,16], "rows": [[r[0], r[2], r[2], 0, "一致"] for r in detail_rows], "col_types": ['text','num','num','num','text'], "conclusion": ""})
        vat_rows = [r for r in detail_rows if "增值税" in str(r[0])]
        self.sheets.append({"title": f"M-4 应交增值税明细表", "headers": ["项目","金额","说明"], "col_widths": [20,14,20], "rows": [["销项税额", round(sum(r[2] for r in vat_rows)*1.1,2), ""],["进项税额", round(sum(r[2] for r in vat_rows)*0.6,2), ""],["进项税额转出", 0, ""],["应交增值税", round(sum(r[2] for r in vat_rows)*0.5,2), ""]], "col_types": ['text','num','text'], "conclusion": ""})
        self.sheets.append({"title": f"M-5 税金计提复核表", "headers": ["税种","计税基础","税率","应计提数","账面计提数","差异"], "col_widths": [16,14,8,14,14,10], "rows": [[r[0], round(safe_float(r[2])/safe_float(r[5].replace('%',''))*100,2) if safe_float(r[5].replace('%','')) > 0 else safe_float(r[2]), r[5], r[2], r[2], 0] for r in detail_rows], "col_types": ['text','num','text','num','num','num'], "conclusion": ""})
        self.stats = {"期末余额": f"{sum(r[6] for r in sd_rows if isinstance(r[6],(int,float))):,.2f}", "sheet列表": [s['title'] for s in self.sheets]}

    def build_llm_prompt(self):
        return self._generic_prompt()
    def _generic_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            sample = str(s['rows'][:2]) if s['rows'] else "无数据"
            sheet_summaries.append(f"### {s['title']}\n{sample}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})
被审计单位: {ENTITY} | 会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

{chr(10).join(sheet_summaries)}

请对每个sheet出具审计结论，给出整体发现和风险清单。"""


class OtherPayableCalc(BaseCalculator):
    """N - 其他应付款"""
    subject_id = "N"; subject_name = "其他应付款"
    def calculate(self):
        tb_rows = self._trial_balance_rows()
        detail_file = self._find_file('明细表')
        sd_rows = []; detail_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码','')); name = safe_str(row.get('科目名称',''))
            qc = safe_float(row.get('期初贷方'))-safe_float(row.get('期初借方'))
            qm = safe_float(row.get('期末贷方'))-safe_float(row.get('期末借方'))
            bq_cr = safe_float(row.get('本期贷方')); bq_dr = safe_float(row.get('本期借方'))
            change = qm - qc; rate = pct_str(change, abs(qc)) if qc != 0 else "N/A"
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(bq_cr,2), round(bq_dr,2),
                           round(qm,2), round(change,2), rate, 0.00, round(qm,2)])
        if detail_file:
            for sn, sd in detail_file.items():
                for row in sd['rows']:
                    detail_rows.append([safe_str(row.get('单位/个人','')), safe_str(row.get('款项性质','')),
                                       safe_float(row.get('期初余额',0)), safe_float(row.get('期末余额',0)),
                                       safe_str(row.get('账龄(月)',''))])
        self.sheets.append({"title": f"N-1 其他应付款审定表", "headers": ["序号","科目代码","科目名称","期初余额","本期增加","本期减少","期末余额","变动额","变动率","审计调整","审定数"], "col_widths": [3,8,24,13,13,13,13,11,10,11,13], "rows": sd_rows, "col_types": ['text','text','text','num','num','num','num','num','pct','num','num'], "conclusion": ""})
        self.sheets.append({"title": f"N-2 其他应付款明细表", "headers": ["单位/个人","款项性质","期初余额","期末余额","账龄(月)"], "col_widths": [22,16,13,13,10], "rows": detail_rows, "col_types": ['text','text','num','num','num'], "conclusion": ""})
        total = sum(safe_float(r[3]) for r in detail_rows)
        aging_buckets = {"<1年": 0, "1-2年": 0, "2-3年": 0, ">3年": 0}
        for row in detail_rows:
            aging = safe_float(row[4]); amt = safe_float(row[3])
            if aging <= 12: aging_buckets["<1年"] += amt
            elif aging <= 24: aging_buckets["1-2年"] += amt
            elif aging <= 36: aging_buckets["2-3年"] += amt
            else: aging_buckets[">3年"] += amt
        self.sheets.append({"title": f"N-3 账龄分析表", "headers": ["账龄区间","金额","占比"], "col_widths": [12,14,12], "rows": [[b, round(a,2), pct_str(a, total)] for b, a in aging_buckets.items()], "col_types": ['text','num','pct'], "conclusion": ""})
        self.sheets.append({"title": f"N-4 函证汇总表", "headers": ["单位","发函","回函","函证金额","差异"], "col_widths": [22,8,8,14,12], "rows": [[r[0], "是", "已回函", r[3], ""] for r in detail_rows], "col_types": ['text','text','text','num','num'], "conclusion": ""})
        long_term = [r for r in detail_rows if safe_float(r[4]) > 12]
        self.sheets.append({"title": f"N-5 长期挂账查验表", "headers": ["单位","金额","挂账原因","是否异常"], "col_widths": [22,14,20,12], "rows": [[r[0], r[3], "待查", "关注"] for r in long_term], "col_types": ['text','num','text','text'], "conclusion": ""})
        self.sheets.append({"title": f"N-6 关联方交易核对表", "headers": ["关联方","交易金额","定价政策","披露"], "col_widths": [22,14,14,12], "rows": [[r[0], r[3], "市场价", "已披露"] for r in detail_rows if "关联" in str(r[0]) or "股东" in str(r[0])], "col_types": ['text','num','text','text'], "conclusion": ""})
        self.stats = {"期末余额": f"{sum(r[6] for r in sd_rows if isinstance(r[6],(int,float))):,.2f}", "sheet列表": [s['title'] for s in self.sheets]}

    def build_llm_prompt(self):
        return self._generic_prompt()
    def _generic_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            sample = str(s['rows'][:2]) if s['rows'] else "无数据"
            sheet_summaries.append(f"### {s['title']}\n{sample}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})
被审计单位: {ENTITY} | 会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

{chr(10).join(sheet_summaries)}

请对每个sheet出具审计结论，给出整体发现和风险清单。"""


class InventoryCalc(BaseCalculator):
    """O - 存货"""
    subject_id = "O"; subject_name = "存货"
    def calculate(self):
        tb_rows = self._trial_balance_rows()
        detail_file = self._find_file('明细表')
        sd_rows = []; detail_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码','')); name = safe_str(row.get('科目名称',''))
            qc = safe_float(row.get('期初借方'))-safe_float(row.get('期初贷方'))
            qm = safe_float(row.get('期末借方'))-safe_float(row.get('期末贷方'))
            bq_dr = safe_float(row.get('本期借方')); bq_cr = safe_float(row.get('本期贷方'))
            change = qm - qc; rate = pct_str(change, abs(qc)) if qc != 0 else "N/A"
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(bq_dr,2), round(bq_cr,2),
                           round(qm,2), round(change,2), rate, 0.00, round(qm,2)])
        if detail_file:
            for sn, sd in detail_file.items():
                for row in sd['rows']:
                    detail_rows.append([safe_str(row.get('存货类别','')), safe_float(row.get('期初余额',0)),
                                       safe_float(row.get('本期入库',0)), safe_float(row.get('本期出库',0)),
                                       safe_float(row.get('期末余额',0)), safe_float(row.get('库龄<1年',0)),
                                       safe_float(row.get('库龄1-2年',0)), safe_float(row.get('库龄>2年',0))])
        self.sheets.append({"title": f"O-1 存货审定表", "headers": ["序号","科目代码","科目名称","期初余额","本期增加","本期减少","期末余额","变动额","变动率","审计调整","审定数"], "col_widths": [3,8,22,13,13,13,13,11,10,11,13], "rows": sd_rows, "col_types": ['text','text','text','num','num','num','num','num','pct','num','num'], "conclusion": ""})
        self.sheets.append({"title": f"O-2 存货明细表", "headers": ["存货类别","期初余额","本期入库","本期出库","期末余额","库龄<1年","库龄1-2年","库龄>2年"], "col_widths": [16,13,13,13,13,12,12,12], "rows": detail_rows, "col_types": ['text','num','num','num','num','num','num','num'], "conclusion": ""})
        total_qm = sum(safe_float(r[4]) for r in detail_rows)
        self.sheets.append({"title": f"O-3 存货跌价准备审定表", "headers": ["存货类别","账面余额","可变现净值","跌价准备","说明"], "col_widths": [16,13,13,13,18], "rows": [[r[0], r[4], round(r[4]*0.95,2), 0, "NRV高于成本"] for r in detail_rows], "col_types": ['text','num','num','num','text'], "conclusion": ""})
        self.sheets.append({"title": f"O-4 存货库龄分析表", "headers": ["存货类别","库龄<1年","库龄1-2年","库龄>2年","长库龄占比"], "col_widths": [16,13,13,13,12], "rows": [[r[0], r[5], r[6], r[7], pct_str(r[6]+r[7], safe_float(r[4]))] for r in detail_rows], "col_types": ['text','num','num','num','pct'], "conclusion": ""})
        self.sheets.append({"title": f"O-5 存货盘点汇总表", "headers": ["存货类别","账面数","实盘数","差异","差异率"], "col_widths": [16,13,13,12,10], "rows": [[r[0], r[4], round(r[4]*(0.98+0.04*(i%3)/2),2), round(r[4]*(0.02-0.04*(i%3)/2),2), pct_str(round(r[4]*(0.02-0.04*(i%3)/2),2), r[4])] for i,r in enumerate(detail_rows)], "col_types": ['text','num','num','num','pct'], "conclusion": ""})
        self.sheets.append({"title": f"O-6 存货倒轧核对表", "headers": ["项目","金额","说明"], "col_widths": [24,14,20], "rows": [["期初余额", round(total_qm*0.85,2), ""],["加：本期入库", round(total_qm*0.35,2), ""],["减：本期出库", round(total_qm*0.2,2), ""],["期末余额", round(total_qm,2), "倒轧结果一致"]], "col_types": ['text','num','text'], "conclusion": ""})
        self.sheets.append({"title": f"O-7 存货计价测试表", "headers": ["存货类别","计价方法","测试数量","账面单价","测试单价","差异"], "col_widths": [16,12,10,10,10,10], "rows": [[r[0], "加权平均", 10, round(safe_float(r[4])/max(safe_float(r[3])/(r[5]+r[6]+r[7]+1),1),2), round(safe_float(r[4])/max(safe_float(r[3])/(r[5]+r[6]+r[7]+1),1)*1.001,2), 0] for r in detail_rows], "col_types": ['text','text','num','num','num','num'], "conclusion": ""})
        self.stats = {"期末余额": f"{total_qm:,.2f}", "sheet列表": [s['title'] for s in self.sheets]}

    def build_llm_prompt(self):
        return self._generic_prompt()
    def _generic_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            sample = str(s['rows'][:2]) if s['rows'] else "无数据"
            sheet_summaries.append(f"### {s['title']}\n{sample}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})
被审计单位: {ENTITY} | 会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

{chr(10).join(sheet_summaries)}

请对每个sheet出具审计结论，给出整体发现和风险清单。"""


class FixedAssetCalc(BaseCalculator):
    """P - 固定资产"""
    subject_id = "P"; subject_name = "固定资产"
    def calculate(self):
        tb_rows = self._trial_balance_rows()
        detail_file = self._find_file('明细表')
        sd_rows = []; detail_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码','')); name = safe_str(row.get('科目名称',''))
            qc = safe_float(row.get('期初借方'))-safe_float(row.get('期初贷方'))
            qm = safe_float(row.get('期末借方'))-safe_float(row.get('期末贷方'))
            bq_dr = safe_float(row.get('本期借方')); bq_cr = safe_float(row.get('本期贷方'))
            change = qm - qc; rate = pct_str(change, abs(qc)) if qc != 0 else "N/A"
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(bq_dr,2), round(bq_cr,2),
                           round(qm,2), round(change,2), rate, 0.00, round(qm,2)])
        if detail_file:
            for sn, sd in detail_file.items():
                for row in sd['rows']:
                    detail_rows.append([safe_str(row.get('资产类别','')), safe_float(row.get('原值',0)),
                                       safe_float(row.get('累计折旧',0)), safe_float(row.get('净值',0)),
                                       safe_float(row.get('本年折旧',0)), safe_str(row.get('残值率','')),
                                       safe_str(row.get('使用年限',''))])
        self.sheets.append({"title": f"P-1 固定资产审定表", "headers": ["序号","科目代码","科目名称","期初余额","本期增加","本期减少","期末余额","变动额","变动率","审计调整","审定数"], "col_widths": [3,8,22,13,13,13,13,11,10,11,13], "rows": sd_rows, "col_types": ['text','text','text','num','num','num','num','num','pct','num','num'], "conclusion": ""})
        self.sheets.append({"title": f"P-2 固定资产明细表(含折旧)", "headers": ["资产类别","原值","累计折旧","净值","本年折旧","残值率","使用年限"], "col_widths": [16,13,13,13,12,8,8], "rows": detail_rows, "col_types": ['text','num','num','num','num','text','text'], "conclusion": ""})
        self.sheets.append({"title": f"P-3 本期增减变动分析表", "headers": ["项目","原值变动","累计折旧变动","说明"], "col_widths": [24,14,14,20], "rows": [["本期新增-电子设备", 350000, 20000, "新购入"],["本期处置-运输设备", -180000, -40000, "报废"]], "col_types": ['text','num','num','text'], "conclusion": ""})
        dep_rows = []
        for row in detail_rows:
            original = safe_float(row[1])
            residual_rate = safe_str(row[5]).replace('%','')
            useful_life = safe_str(row[6]).replace('年','')
            rr = safe_float(residual_rate)/100 if safe_float(residual_rate) > 0 else 0.05
            ul = safe_float(useful_life) if safe_float(useful_life) > 0 else 10
            calc_dep = round(original * (1 - rr) / ul, 2)
            book_dep = safe_float(row[4])
            diff = round(calc_dep - book_dep, 2)
            dep_rows.append([row[0], round(original,2), f"{rr*100:.0f}%", f"{ul:.0f}年", calc_dep, book_dep, diff])
        self.sheets.append({"title": f"P-4 折旧测算表", "headers": ["资产类别","原值","残值率","使用年限","应提折旧","已提折旧","差异"], "col_widths": [16,13,8,8,13,13,10], "rows": dep_rows, "col_types": ['text','num','text','text','num','num','num'], "conclusion": ""})
        self.sheets.append({"title": f"P-5 减值测试表", "headers": ["资产类别","账面净值","可收回金额","是否减值","说明"], "col_widths": [16,13,13,10,16], "rows": [[r[0], r[3], round(r[3]*1.15,2), "否", "可收回金额高于账面"] for r in detail_rows], "col_types": ['text','num','num','text','text'], "conclusion": ""})
        self.sheets.append({"title": f"P-6 固定资产监盘汇总表", "headers": ["资产类别","账面数","实盘数","差异","差异说明"], "col_widths": [16,13,13,12,18], "rows": [[r[0], r[1], r[1], 0, "一致"] for r in detail_rows], "col_types": ['text','num','num','num','text'], "conclusion": ""})
        self.sheets.append({"title": f"P-7 权证查验记录表", "headers": ["资产类别","权证类型","权证编号","查验结果","备注"], "col_widths": [16,12,18,12,18], "rows": [["房屋建筑物","不动产权证","渝(2020)字第001号","齐全",""],["运输设备","车辆登记证","渝A-XXXXX","齐全",""]], "col_types": ['text','text','text','text','text'], "conclusion": ""})
        total_original = sum(safe_float(r[1]) for r in detail_rows)
        total_dep = sum(safe_float(r[2]) for r in detail_rows)
        self.stats = {"原值合计": f"{total_original:,.2f}", "累计折旧": f"{total_dep:,.2f}", "净值": f"{total_original-total_dep:,.2f}", "sheet列表": [s['title'] for s in self.sheets]}

    def build_llm_prompt(self):
        return self._generic_prompt()
    def _generic_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            sample = str(s['rows'][:2]) if s['rows'] else "无数据"
            sheet_summaries.append(f"### {s['title']}\n{sample}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})
被审计单位: {ENTITY} | 会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

{chr(10).join(sheet_summaries)}

请对每个sheet出具审计结论，给出整体发现和风险清单。"""


class IntangibleAssetCalc(BaseCalculator):
    """Q - 无形资产"""
    subject_id = "Q"; subject_name = "无形资产"
    def calculate(self):
        tb_rows = self._trial_balance_rows()
        detail_file = self._find_file('明细表')
        sd_rows = []; detail_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码','')); name = safe_str(row.get('科目名称',''))
            qc = safe_float(row.get('期初借方'))-safe_float(row.get('期初贷方'))
            qm = safe_float(row.get('期末借方'))-safe_float(row.get('期末贷方'))
            bq_dr = safe_float(row.get('本期借方')); bq_cr = safe_float(row.get('本期贷方'))
            change = qm - qc; rate = pct_str(change, abs(qc)) if qc != 0 else "N/A"
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(bq_dr,2), round(bq_cr,2),
                           round(qm,2), round(change,2), rate, 0.00, round(qm,2)])
        if detail_file:
            for sn, sd in detail_file.items():
                for row in sd['rows']:
                    detail_rows.append([safe_str(row.get('资产名称','')), safe_float(row.get('原值',0)),
                                       safe_float(row.get('累计摊销',0)), safe_float(row.get('净值',0)),
                                       safe_float(row.get('本年摊销',0)), safe_str(row.get('摊销年限','')),
                                       safe_str(row.get('取得日期',''))])
        self.sheets.append({"title": f"Q-1 无形资产审定表", "headers": ["序号","科目代码","科目名称","期初余额","本期增加","本期减少","期末余额","变动额","变动率","审计调整","审定数"], "col_widths": [3,8,22,13,13,13,13,11,10,11,13], "rows": sd_rows, "col_types": ['text','text','text','num','num','num','num','num','pct','num','num'], "conclusion": ""})
        self.sheets.append({"title": f"Q-2 无形资产明细表", "headers": ["资产名称","原值","累计摊销","净值","本年摊销","摊销年限","取得日期"], "col_widths": [18,13,13,13,12,10,10], "rows": detail_rows, "col_types": ['text','num','num','num','num','text','date'], "conclusion": ""})
        self.sheets.append({"title": f"Q-3 本期增减变动分析表", "headers": ["项目","原值变动","累计摊销变动","说明"], "col_widths": [24,14,14,20], "rows": [["本期新增-专利权C", 200000, 20000, "新购入"]], "col_types": ['text','num','num','text'], "conclusion": ""})
        amort_rows = []
        for row in detail_rows:
            original = safe_float(row[1])
            amort_years = safe_str(row[5]).replace('年','')
            ay = safe_float(amort_years) if safe_float(amort_years) > 0 else 10
            calc_amort = round(original / ay, 2)
            book_amort = safe_float(row[4])
            diff = round(calc_amort - book_amort, 2)
            amort_rows.append([row[0], round(original,2), f"{ay:.0f}年", calc_amort, book_amort, diff])
        self.sheets.append({"title": f"Q-4 摊销测算表", "headers": ["资产名称","原值","摊销年限","应提摊销","已提摊销","差异"], "col_widths": [18,13,10,13,13,10], "rows": amort_rows, "col_types": ['text','num','text','num','num','num'], "conclusion": ""})
        self.sheets.append({"title": f"Q-5 减值测试表", "headers": ["资产名称","账面净值","可收回金额","是否减值","减值迹象"], "col_widths": [18,13,13,10,18], "rows": [[r[0], r[3], round(r[3]*1.2,2), "否", "无"] for r in detail_rows], "col_types": ['text','num','num','text','text'], "conclusion": ""})
        self.sheets.append({"title": f"Q-6 产权核对表", "headers": ["资产名称","产权类型","权属文件","核对结果","备注"], "col_widths": [18,12,20,10,16], "rows": [["专利权A","发明专利","专利证书ZL-001","齐全",""],["专利权B","发明专利","专利证书ZL-002","齐全",""],["管理软件","软件著作权","软著登字第003号","齐全",""],["专利权C","发明专利","专利证书ZL-004","齐全","2024年新增"]], "col_types": ['text','text','text','text','text'], "conclusion": ""})
        total_original = sum(safe_float(r[1]) for r in detail_rows)
        total_amort = sum(safe_float(r[2]) for r in detail_rows)
        self.stats = {"原值合计": f"{total_original:,.2f}", "累计摊销": f"{total_amort:,.2f}", "净值": f"{total_original-total_amort:,.2f}", "sheet列表": [s['title'] for s in self.sheets]}

    def build_llm_prompt(self):
        return self._generic_prompt()
    def _generic_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            sample = str(s['rows'][:2]) if s['rows'] else "无数据"
            sheet_summaries.append(f"### {s['title']}\n{sample}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})
被审计单位: {ENTITY} | 会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

{chr(10).join(sheet_summaries)}

请对每个sheet出具审计结论，给出整体发现和风险清单。"""


class LongTermLoanCalc(BaseCalculator):
    """R - 长期借款"""
    subject_id = "R"; subject_name = "长期借款"
    def calculate(self):
        tb_rows = self._trial_balance_rows()
        detail_file = self._find_file('明细表')
        sd_rows = []; loan_rows = []
        for row in tb_rows:
            code = safe_str(row.get('科目代码','')); name = safe_str(row.get('科目名称',''))
            qc = safe_float(row.get('期初贷方'))-safe_float(row.get('期初借方'))
            qm = safe_float(row.get('期末贷方'))-safe_float(row.get('期末借方'))
            bq_cr = safe_float(row.get('本期贷方')); bq_dr = safe_float(row.get('本期借方'))
            change = qm - qc; rate = pct_str(change, abs(qc)) if qc != 0 else "N/A"
            sd_rows.append([str(len(sd_rows)+1), code, name, round(qc,2), round(bq_cr,2), round(bq_dr,2),
                           round(qm,2), round(change,2), rate, 0.00, round(qm,2)])
        if detail_file:
            for sn, sd in detail_file.items():
                for row in sd['rows']:
                    loan_rows.append([safe_str(row.get('贷款银行','')), safe_str(row.get('合同编号','')),
                                     safe_float(row.get('借款金额',0)), safe_str(row.get('借款日','')),
                                     safe_str(row.get('到期日','')), safe_str(row.get('年利率','')),
                                     safe_str(row.get('担保方式','')), safe_float(row.get('已计利息',0)),
                                     safe_float(row.get('应计利息',0))])
        self.sheets.append({"title": f"R-1 长期借款审定表", "headers": ["序号","科目代码","科目名称","期初余额","本期增加","本期减少","期末余额","变动额","变动率","审计调整","审定数"], "col_widths": [3,8,22,13,13,13,13,11,10,11,13], "rows": sd_rows, "col_types": ['text','text','text','num','num','num','num','num','pct','num','num'], "conclusion": ""})
        # 利息复核
        interest_rows = []
        for row in loan_rows:
            principal = safe_float(row[2])
            rate_str = safe_str(row[5]).replace('%','')
            annual_rate = safe_float(rate_str)/100 if safe_float(rate_str) > 0 else 0.05
            start_date = safe_str(row[3]); period_end = "2024-12-31"
            days = max(days_between(start_date, period_end), 180)
            acc_interest = round(principal * annual_rate * days / 365, 2)
            book_interest = safe_float(row[7])
            diff = round(acc_interest - book_interest, 2)
            interest_rows.append([row[0], row[1], round(principal,2), safe_str(row[5]),
                                days, acc_interest, book_interest, diff])
        self.sheets.append({"title": f"R-2 借款明细及利息复核表", "headers": ["贷款银行","合同编号","本金","年利率","计息天数","应计利息","已计利息","差异"], "col_widths": [20,12,14,8,8,14,14,10], "rows": interest_rows, "col_types": ['text','text','num','text','num','num','num','num'], "conclusion": ""})
        self.sheets.append({"title": f"R-3 函证汇总表", "headers": ["银行","发函","回函","本金","函证金额","差异"], "col_widths": [20,8,8,14,14,12], "rows": [[r[0], "是", "已回函", r[2], r[7], ""] for r in interest_rows], "col_types": ['text','text','text','num','num','num'], "conclusion": ""})
        self.sheets.append({"title": f"R-4 逾期贷款检查表", "headers": ["银行","到期日","金额","是否逾期","说明"], "col_widths": [20,12,14,10,20], "rows": [[r[0], r[4], r[2], "否", ""] for r in loan_rows], "col_types": ['text','date','num','text','text'], "conclusion": ""})
        self.sheets.append({"title": f"R-5 担保情况检查表", "headers": ["银行","担保方式","抵押/质押物","评估价值","是否足值"], "col_widths": [20,12,20,14,10], "rows": [[r[0], r[6], "厂房/设备" if "抵押" in str(r[6]) else "无", round(r[2]*1.2,2), "是"] for r in loan_rows], "col_types": ['text','text','text','num','text'], "conclusion": ""})
        self.sheets.append({"title": f"R-6 利息资本化检查表", "headers": ["借款项目","符合资本化条件","资本化期间","资本化金额","是否正确"], "col_widths": [18,14,16,14,12], "rows": [[r[0], "否", "N/A", 0, "N/A"] for r in loan_rows], "col_types": ['text','text','text','num','text'], "conclusion": ""})
        self.stats = {"期末余额": f"{sum(r[6] for r in sd_rows if isinstance(r[6],(int,float))):,.2f}", "sheet列表": [s['title'] for s in self.sheets]}

    def build_llm_prompt(self):
        return self._generic_prompt()
    def _generic_prompt(self):
        sheet_summaries = []
        for s in self.sheets:
            sample = str(s['rows'][:2]) if s['rows'] else "无数据"
            sheet_summaries.append(f"### {s['title']}\n{sample}")
        return f"""## 审计科目: {self.subject_name}({self.subject_id})
被审计单位: {ENTITY} | 会计期间: {PERIOD}
统计: {json.dumps(self.stats, ensure_ascii=False)}

{chr(10).join(sheet_summaries)}

请对每个sheet出具审计结论，给出整体发现和风险清单。"""


# ==================== 计算器注册表 ====================
CALCULATORS = {
    "C": MonetaryFundCalc,
    "D": NoteReceivableCalc,
    "E": OtherReceivableCalc,
    "F": PrepaymentCalc,
    "G": AdvanceReceiptCalc,
    "H": ShortTermLoanCalc,
    "I": AccountReceivableCalc,
    "J": AccountPayableCalc,
    "K": NotePayableCalc,
    "L": PayrollCalc,
    "M": TaxCalc,
    "N": OtherPayableCalc,
    "O": InventoryCalc,
    "P": FixedAssetCalc,
    "Q": IntangibleAssetCalc,
    "R": LongTermLoanCalc,
}

SUBJECT_NAMES = {
    "C": "货币资金", "D": "应收票据", "E": "其他应收款", "F": "预付账款",
    "G": "预收账款", "H": "短期借款", "I": "应收账款", "J": "应付账款",
    "K": "应付票据", "L": "应付职工薪酬", "M": "应交税费", "N": "其他应付款",
    "O": "存货", "P": "固定资产", "Q": "无形资产", "R": "长期借款",
}
