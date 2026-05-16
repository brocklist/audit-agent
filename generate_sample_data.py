"""
生成货币资金审计测试数据
输出: 科目余额表.xlsx + 明细账.xlsx
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
import random

BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
FONT_HEAD = Font(name="微软雅黑", bold=True, size=10)
FONT_NORM = Font(name="微软雅黑", size=10)
FILL_HEAD = PatternFill("solid", fgColor="D6E4F0")

def style_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row, c)
        cell.font = FONT_HEAD; cell.fill = FILL_HEAD; cell.alignment = Alignment(horizontal="center"); cell.border = BORDER

def style_data(ws, row, ncols):
    for c in range(1, ncols+1):
        ws.cell(row, c).font = FONT_NORM; ws.cell(row, c).border = BORDER

# ============================================================
# 科目余额表
# ============================================================
wb1 = openpyxl.Workbook()
ws = wb1.active
ws.title = "科目余额表"

headers = ["科目代码", "科目名称", "期初借方余额", "期初贷方余额", "本期借方发生额", "本期贷方发生额", "期末借方余额", "期末贷方余额"]
for c, h in enumerate(headers, 1):
    ws.cell(1, c, h)
style_header(ws, 1, len(headers))

data = [
    # 货币资金类
    ["1001", "库存现金", 50000, 0, 120000, 115000, 55000, 0],
    ["1002", "银行存款-工行渝北支行", 2500000, 0, 8650000, 8320000, 2830000, 0],
    ["100201", "银行存款-建行两江支行", 1800000, 0, 5200000, 4980000, 2020000, 0],
    ["100202", "银行存款-招行重庆分行", 950000, 0, 3100000, 3200000, 850000, 0],
    ["100203", "银行存款-农行江北支行", 620000, 0, 1850000, 1780000, 690000, 0],
    ["1012", "其他货币资金-保证金", 300000, 0, 0, 0, 300000, 0],
    ["101201", "其他货币资金-银行承兑汇票保证金", 200000, 0, 0, 0, 200000, 0],
    # 非货币资金类（验证智能体能否正确筛选）
    ["1122", "应收账款", 1680000, 0, 38600000, 38350000, 1850000, 0],
    ["112201", "应收账款-湖北双环科技", 1200000, 0, 28000000, 27900000, 1300000, 0],
    ["1221", "其他应收款", 85000, 0, 450000, 420000, 115000, 0],
    ["1403", "原材料", 3200000, 0, 8500000, 7900000, 3800000, 0],
    ["1405", "库存商品", 5100000, 0, 28000000, 27500000, 5600000, 0],
    ["1601", "固定资产", 12500000, 0, 350000, 180000, 12670000, 0],
    ["2202", "应付账款", 0, 2350000, 18000000, 18250000, 0, 2600000],
    ["4001", "实收资本", 0, 10000000, 0, 0, 0, 10000000],
    ["4104", "利润分配-未分配利润", 0, 3500000, 0, 1200000, 0, 4700000],
]

for i, row in enumerate(data, 2):
    for c, v in enumerate(row, 1):
        ws.cell(i, c, v)
    style_data(ws, i, len(headers))
    # 金额列千分位
    for c in range(3, len(headers)+1):
        ws.cell(i, c).number_format = '#,##0.00'

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 30
for col in ['C', 'D', 'E', 'F', 'G', 'H']:
    ws.column_dimensions[col].width = 16

wb1.save(r"D:\audit-agent\科目余额表_测试数据.xlsx")
print("科目余额表 已生成")

# ============================================================
# 明细账（银行存款-工行渝北支行）
# ============================================================
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "银行存款明细账-工行渝北支行"

h2 = ["日期", "凭证号", "摘要", "借方金额", "贷方金额", "余额"]
for c, h in enumerate(h2, 1):
    ws2.cell(1, c, h)
style_header(ws2, 1, len(h2))

base_date = datetime(2024, 1, 1)
transactions = [
    ("2024-01-05", "记-001", "收到重庆XX公司货款", 520000, 0),
    ("2024-01-15", "记-012", "支付供应商款项", 0, 380000),
    ("2024-02-08", "记-023", "收到四川XX公司预付款", 680000, 0),
    ("2024-02-20", "记-035", "支付员工工资", 0, 450000),
    ("2024-03-12", "记-047", "收到货款-湖北双环科技", 850000, 0),
    ("2024-03-25", "记-058", "支付税费", 0, 320000),
    ("2024-04-10", "记-069", "收到应收账款", 720000, 0),
    ("2024-04-22", "记-081", "支付房租", 0, 180000),
    ("2024-05-15", "记-093", "收到XX投资公司投资款", 2000000, 0),
    ("2024-05-28", "记-105", "支付设备采购款", 0, 1200000),
    ("2024-06-10", "记-117", "收到货款", 460000, 0),
    ("2024-06-25", "记-129", "支付工程款", 0, 650000),
    ("2024-07-08", "记-141", "收到政府补贴", 500000, 0),
    ("2024-07-20", "记-153", "支付供应商尾款", 0, 290000),
    ("2024-08-12", "记-165", "收到货款", 380000, 0),
    ("2024-08-28", "记-177", "支付咨询费", 0, 85000),
    ("2024-09-15", "记-189", "收到货款", 620000, 0),
    ("2024-09-28", "记-201", "支付保证金", 0, 200000),
    ("2024-10-10", "记-213", "收到应收账款", 410000, 0),
    ("2024-10-25", "记-225", "支付运输费", 0, 135000),
    ("2024-11-12", "记-237", "收到货款", 350000, 0),
    ("2024-11-28", "记-249", "支付水电费", 0, 95000),
    ("2024-12-10", "记-261", "收到货款", 280000, 0),
    ("2024-12-20", "记-273", "支付年终奖", 0, 580000),
    ("2024-12-31", "记-285", "收到银行利息", 12000, 0),
    ("2024-12-31", "记-286", "大额转账至建行", 0, 500000),  # 大额异常转账标记
]

bal = 2500000
for i, (date, voucher, summary, debit, credit) in enumerate(transactions, 2):
    ws2.cell(2, 1, value=date)
    ws2.cell(i, 1, value=date)
    ws2.cell(i, 2, value=voucher)
    ws2.cell(i, 3, value=summary)
    ws2.cell(i, 4, value=debit if debit else "")
    ws2.cell(i, 5, value=credit if credit else "")
    bal = bal + (debit or 0) - (credit or 0)
    ws2.cell(i, 6, value=bal)
    style_data(ws2, i, len(h2))
    for c in range(4, 7):
        ws2.cell(i, c).number_format = '#,##0.00'

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 28
for col in ['D', 'E', 'F']:
    ws2.column_dimensions[col].width = 18

wb2.save(r"D:\audit-agent\明细账_测试数据.xlsx")
print("明细账 已生成")

# ============================================================
# 银行对账单（工行渝北支行）
# ============================================================
wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.title = "银行对账单-工行渝北支行"

h3 = ["日期", "摘要", "收入", "支出", "余额"]
for c, h in enumerate(h3, 1):
    ws3.cell(1, c, h)
style_header(ws3, 1, len(h3))

bank_statement = [
    ("2024-01-05", "货款入账", 520000, 0),
    ("2024-01-15", "支付货款", 0, 380000),
    ("2024-02-08", "预付款入账", 680000, 0),
    ("2024-02-20", "工资代发", 0, 450000),
    ("2024-03-12", "货款入账", 850000, 0),
    ("2024-03-25", "扣税", 0, 320000),
    ("2024-04-10", "货款入账", 720000, 0),
    ("2024-04-22", "转账支出", 0, 180000),
    ("2024-05-15", "投资款入账", 2000000, 0),
    ("2024-05-28", "设备款支出", 0, 1200000),
    ("2024-06-10", "货款入账", 460000, 0),
    ("2024-06-25", "工程款支出", 0, 650000),
    ("2024-07-08", "补贴入账", 500000, 0),
    ("2024-07-20", "尾款支出", 0, 290000),
    ("2024-08-12", "货款入账", 380000, 0),
    ("2024-08-28", "咨询费支出", 0, 85000),
    ("2024-09-15", "货款入账", 620000, 0),
    ("2024-09-28", "保证金支出", 0, 200000),
    ("2024-10-10", "货款入账", 410000, 0),
    ("2024-10-25", "运输费支出", 0, 135000),
    ("2024-11-12", "货款入账", 350000, 0),
    ("2024-11-28", "水电费支出", 0, 95000),
    ("2024-12-10", "货款入账", 280000, 0),
    ("2024-12-20", "奖金发放", 0, 580000),
    ("2024-12-31", "利息入账", 12000, 0),
    ("2024-12-31", "跨行转出", 0, 500000),
    # 银行已收企业未收
    ("2024-12-31", "银行存款利息（企业未记账）", 3500, 0),
    # 银行已付企业未付
    ("2024-12-31", "账户管理费（企业未记账）", 0, 1200),
]

bal = 2500000
for i, (date, summary, income, expense) in enumerate(bank_statement, 2):
    ws3.cell(i, 1, value=date)
    ws3.cell(i, 2, value=summary)
    ws3.cell(i, 3, value=income if income else "")
    ws3.cell(i, 4, value=expense if expense else "")
    bal = bal + income - expense
    ws3.cell(i, 5, value=bal)
    style_data(ws3, i, len(h3))
    for c in range(3, 6):
        ws3.cell(i, c).number_format = '#,##0.00'

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 28
for col in ['C', 'D', 'E']:
    ws3.column_dimensions[col].width = 18

wb3.save(r"D:\audit-agent\银行对账单_测试数据.xlsx")
print("银行对账单 已生成")
print("\n测试数据全部生成完毕:")
print("  D:\\audit-agent\\科目余额表_测试数据.xlsx")
print("  D:\\audit-agent\\明细账_测试数据.xlsx")
print("  D:\\audit-agent\\银行对账单_测试数据.xlsx")
